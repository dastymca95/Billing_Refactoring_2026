from __future__ import annotations

import csv
import json
import math
import re
import shutil
import sys
import time
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from utils.dropbox_uploader import DropboxUploader, build_dropbox_path

PUNCTUAL_DIR = ROOT / "Punctual Process"
INVOICE_DIR = PUNCTUAL_DIR / "Invoices"
OUTPUT_PATH = PUNCTUAL_DIR / "Punctual_Process_ResMan_Import.xlsx"
AUDIT_PATH = PUNCTUAL_DIR / "Punctual_Process_Invoice_Audit.csv"
LINK_CACHE_PATH = PUNCTUAL_DIR / "_dropbox_links.json"
TEMPLATE_PATH = ROOT / "Output" / "Template.xlsx"

TKA = "TKA"
TLA = "TLA"

V_APARTMENTS = "Apartments.com"
V_CHADWELL = "Chadwell Supply"
V_HANDY = "Handy Manny Solutions LLC"
V_HEAVENS = "Heaven's Handywork Lawn & Landscaping LLC"
V_CERV = "Cerv Pest Solutions - Nashville"
V_MCCOY = "The Law office of Jennifer Mccoy"
V_REALPAGE = "RealPage One"
V_SOUTHERN = "Southern Acre Landscaping LLC"
V_VERIFAST = "VeriFast Inc."
V_ZILLOW = "Zillow Rentals"
V_SHERWIN = "Sherwin Williams (Nex-Gen)"
V_TENANT_REFUND = "Mitchel Gutu; Zoy Davis"


@dataclass
class SourceDoc:
    filename: str
    vendor: str
    invoice_date: date


def parse_date(value: str) -> date:
    value = value.strip()
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%B %d, %Y", "%b %d, %Y"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Unsupported date: {value}")


def iso(d: date | None) -> str:
    return d.isoformat() if d else ""


def add_month(d: date) -> date:
    month = d.month + 1
    year = d.year + (month - 1) // 12
    month = ((month - 1) % 12) + 1
    last_day = [
        31,
        29 if year % 4 == 0 and (year % 100 != 0 or year % 400 == 0) else 28,
        31,
        30,
        31,
        30,
        31,
        31,
        30,
        31,
        30,
        31,
    ][month - 1]
    return date(year, month, min(d.day, last_day))


def money(value: float) -> float:
    return round(float(value) + 0.0000001, 2)


def month_label(d: date) -> str:
    return d.strftime("%b-%y")


def period_label(start: date, end: date) -> str:
    return f"{start.strftime('%m/%d/%y')}-{end.strftime('%m/%d/%y')}"


def allocate_tax(lines: list[dict[str, Any]], tax_total: float) -> list[dict[str, Any]]:
    taxable = [line for line in lines if line.get("taxable")]
    if not taxable or abs(tax_total) < 0.005:
        for line in lines:
            line["amount"] = money(line["base"])
        return lines
    taxable_total = sum(float(line["base"]) for line in taxable)
    allocated = 0.0
    for line in lines:
        base = float(line["base"])
        if not line.get("taxable"):
            line["amount"] = money(base)
            continue
        if line is taxable[-1]:
            tax = money(tax_total - allocated)
        else:
            tax = money(tax_total * base / taxable_total)
            allocated += tax
        line["amount"] = money(base + tax)
    return lines


def load_link_cache() -> dict[str, str]:
    if LINK_CACHE_PATH.exists():
        return json.loads(LINK_CACHE_PATH.read_text(encoding="utf-8"))
    return {}


def save_link_cache(cache: dict[str, str]) -> None:
    LINK_CACHE_PATH.write_text(json.dumps(cache, indent=2, sort_keys=True), encoding="utf-8")


def upload_links(source_docs: dict[str, SourceDoc]) -> dict[str, str]:
    uploader = DropboxUploader.from_env()
    if not uploader.is_configured:
        raise RuntimeError("Dropbox is not configured; Document Url cannot be completed.")
    cache = load_link_cache()
    links: dict[str, str] = {}
    for key, doc in sorted(source_docs.items()):
        local_path = INVOICE_DIR / doc.filename
        if not local_path.is_file():
            raise FileNotFoundError(local_path)
        cache_key = f"{doc.filename}|{local_path.stat().st_size}|{int(local_path.stat().st_mtime)}"
        if cache_key in cache and cache[cache_key].startswith("https://"):
            links[key] = cache[cache_key]
            continue
        dropbox_path = build_dropbox_path(
            base_folder=uploader.base_folder,
            vendor_name=f"Punctual Process/{doc.vendor}",
            billing_date=datetime.combine(doc.invoice_date, datetime.min.time()),
            filename=doc.filename,
        )
        result = None
        for attempt in range(1, 4):
            result = uploader.upload(local_path=local_path, dropbox_path=dropbox_path, overwrite=True)
            if result.success:
                break
            time.sleep(1.5 * attempt)
        if result is None or not result.success:
            msg = result.error_message if result else "unknown Dropbox error"
            raise RuntimeError(f"Dropbox upload failed for {doc.filename}: {msg}")
        cache[cache_key] = result.shared_link
        links[key] = result.shared_link
        save_link_cache(cache)
    return links


def make_row(
    *,
    invoice_number: str,
    invoice_date: date,
    vendor: str,
    invoice_description: str,
    line_item_number: int,
    property_abbreviation: str,
    gl_account: str,
    line_item_description: str,
    amount: float,
    due_date: date,
    source_key: str,
    document_url: str,
    location: str = "",
    quantity: float | int | str = "",
    unit_price: float | int | str = "",
    received_date: date | None = None,
    reference_number: str = "",
) -> dict[str, Any]:
    return {
        "Invoice Number": invoice_number,
        "Bill or Credit": "Bill",
        "Invoice Date": iso(invoice_date),
        "Accounting Date": iso(invoice_date),
        "Vendor": vendor,
        "Invoice Description": invoice_description,
        "Line Item Number": line_item_number,
        "Property Abbreviation": property_abbreviation,
        "Location": location,
        "GL Account": gl_account,
        "Line Item Description": line_item_description,
        "Amount": money(amount),
        "Expense Type": "General",
        "Is Replacement Reserve": "False",
        "Payment Date": "",
        "Reference Number": reference_number,
        "Payment Method": "",
        "Department": "",
        "Due Date": iso(due_date),
        "Quantity": quantity,
        "Unit Price": unit_price,
        "Tax": "",
        "Received Date": iso(received_date or invoice_date),
        "Document Url": document_url,
        "_source_key": source_key,
    }


def add_simple_invoice(
    rows: list[dict[str, Any]],
    *,
    links: dict[str, str],
    source_docs: dict[str, SourceDoc],
    source_key: str,
    invoice_number: str,
    invoice_date: date,
    due_date: date,
    vendor: str,
    prop: str,
    gl: str,
    amount: float,
    invoice_description: str,
    line_description: str,
    location: str = "",
    quantity: float | int | str = 1,
    unit_price: float | int | str | None = None,
    reference_number: str = "",
) -> None:
    rows.append(
        make_row(
            invoice_number=invoice_number,
            invoice_date=invoice_date,
            due_date=due_date,
            vendor=vendor,
            invoice_description=invoice_description,
            line_item_number=1,
            property_abbreviation=prop,
            location=location,
            gl_account=gl,
            line_item_description=line_description,
            amount=amount,
            quantity=quantity,
            unit_price=money(unit_price if unit_price is not None else amount),
            reference_number=reference_number,
            source_key=source_key,
            document_url=links[source_key],
        )
    )


def register(source_docs: dict[str, SourceDoc], key: str, filename: str, vendor: str, invoice_date: str) -> None:
    source_docs[key] = SourceDoc(filename=filename, vendor=vendor, invoice_date=parse_date(invoice_date))


def build_source_docs() -> dict[str, SourceDoc]:
    docs: dict[str, SourceDoc] = {}
    register(docs, "sherwin_7104", "637316092_71040_02252026.pdf", V_SHERWIN, "02/25/2026")
    register(docs, "sherwin_95882", "637316092_95882_09232025.pdf", V_SHERWIN, "09/23/2025")
    register(docs, "apartments_123438269", "Apartments.Com - 123438269.pdf", V_APARTMENTS, "02/02/2026")
    register(docs, "apartments_123606886", "Apartments.Com - 123606886.pdf", V_APARTMENTS, "03/02/2026")
    register(docs, "chadwell_010890135", "Chadwell - 010890135.pdf", V_CHADWELL, "02/26/2026")
    for inv, dt, filename_dt in [
        ("1029", "09/20/2025", "2025-09-20"),
        ("1063", "10/05/2025", "2025-11-04"),
        ("1066", "10/05/2025", "2025-11-04"),
        ("1087", "10/16/2025", "2025-11-15"),
        ("1296", "02/14/2026", "2026-03-16"),
        ("1321", "03/02/2026", "2026-04-01"),
        ("1322", "03/02/2026", "2026-04-01"),
        ("1323", "03/02/2026", "2026-04-01"),
    ]:
        register(docs, f"handy_{inv}", f"Invoice_{inv}_{filename_dt}.pdf", V_HANDY, dt)
    register(docs, "handy_1069", "Handy - 1069.pdf", V_HANDY, "10/05/2025")
    register(docs, "heavens_70242", "kensington_70242_9.3.2025_1.re16fx7gzwwo (1) (002).pdf", V_HEAVENS, "09/03/2025")
    for inv, dt in [
        ("1237104", "08/26/2025"),
        ("1245963", "09/12/2025"),
        ("1261313", "10/12/2025"),
        ("1276179", "11/12/2025"),
        ("1290760", "12/12/2025"),
        ("1303747", "01/12/2026"),
    ]:
        register(docs, f"cerv_{inv}", f"Magna - {inv}.pdf", V_CERV, dt)
    register(docs, "tenant_refund_temp54447402", "Mitchel - TEMP54447402.pdf", V_TENANT_REFUND, "02/24/2026")
    register(docs, "mccoy_40318", "Nex-Gen Management-Trinity Lofts-40318.pdf", V_MCCOY, "04/20/2026")
    for inv, dt in [
        ("25427", "08/23/2025"),
        ("27839", "09/18/2025"),
        ("29074", "10/27/2025"),
        ("31739", "11/26/2025"),
        ("33588", "12/31/2025"),
        ("35963", "02/28/2026"),
    ]:
        register(docs, f"mccoy_{inv}", f"The Kensington Apartments-{inv}.pdf", V_MCCOY, dt)
    for inv, dt in [("35757", "02/28/2026"), ("39786", "03/18/2026")]:
        register(docs, f"mccoy_{inv}", f"Trinity Lofts-{inv}.pdf", V_MCCOY, dt)
    for inv, dt in [
        ("I2504079315", "04/18/2025"),
        ("I2505131914", "05/16/2025"),
        ("I2506139802", "06/18/2025"),
        ("I2507083400", "07/17/2025"),
        ("I2508052387", "08/17/2025"),
        ("I2509133287", "09/17/2025"),
        ("I2510140691", "10/16/2025"),
        ("I2511148820", "11/19/2025"),
        ("I2512076391", "12/17/2025"),
        ("I2601125816", "01/17/2026"),
        ("I2602011770", "02/18/2026"),
    ]:
        register(docs, f"realpage_{inv}", f"Real Page - {inv}.pdf", V_REALPAGE, dt)
    register(docs, "southern_19126", "Southern Acre - 19126.pdf", V_SOUTHERN, "03/01/2026")
    for inv, dt in [("L20959", "12/10/2025"), ("L22797", "01/10/2026"), ("L24470", "02/10/2026")]:
        register(docs, f"verifast_{inv}", f"VeriFast - {inv}.pdf", V_VERIFAST, dt)
    for inv, dt in [("INV34935840", "12/31/2025"), ("INV35306571", "01/31/2026"), ("INV35702646", "02/28/2026")]:
        register(docs, f"zillow_{inv}", f"Zillow - {inv}.pdf", V_ZILLOW, dt)
    return docs


def build_rows(links: dict[str, str], source_docs: dict[str, SourceDoc]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []

    # Sherwin-Williams paint supplies for Trinity Lofts.
    sherwin_7104_lines = allocate_tax(
        [
            {"desc": "PM 200 0 EG extra white paint, 5-gallon", "base": 697.20, "qty": 4, "unit": 174.30, "taxable": True},
            {"desc": "PM 400 0 SG extra white paint, 5-gallon", "base": 277.10, "qty": 2, "unit": 138.55, "taxable": True},
        ],
        94.99,
    )
    inv_date = parse_date("02/25/2026")
    inv_desc = "02/25/26 - Trinity Lofts: Sherwin Williams paint supplies"
    for i, line in enumerate(sherwin_7104_lines, 1):
        rows.append(make_row(
            invoice_number="71040160120226",
            invoice_date=inv_date,
            due_date=add_month(inv_date),
            vendor=V_SHERWIN,
            invoice_description=inv_desc,
            line_item_number=i,
            property_abbreviation=TLA,
            gl_account="6770",
            line_item_description=f"02/25/26 - Trinity Lofts - {line['desc']}",
            amount=line["amount"],
            quantity=line["qty"],
            unit_price=line["unit"],
            source_key="sherwin_7104",
            document_url=links["sherwin_7104"],
            reference_number="6373-1609-2",
        ))
    add_simple_invoice(
        rows,
        links=links,
        source_docs=source_docs,
        source_key="sherwin_95882",
        invoice_number="95882160120925",
        invoice_date=parse_date("09/23/2025"),
        due_date=add_month(parse_date("09/23/2025")),
        vendor=V_SHERWIN,
        prop=TLA,
        gl="6770",
        amount=36.65,
        invoice_description="09/23/25 - Trinity Lofts: Sherwin Williams corrected paint supplies",
        line_description="09/23/25 - Trinity Lofts - Corrected paint supply charge",
        reference_number="6373-1609-2",
    )

    for key, inv, dt, due, period in [
        ("apartments_123438269", "123438269", "02/02/2026", "03/04/2026", "Feb-26"),
        ("apartments_123606886", "123606886", "03/02/2026", "04/01/2026", "Mar-26"),
    ]:
        add_simple_invoice(
            rows,
            links=links,
            source_docs=source_docs,
            source_key=key,
            invoice_number=inv,
            invoice_date=parse_date(dt),
            due_date=parse_date(due),
            vendor=V_APARTMENTS,
            prop=TKA,
            gl="6335",
            amount=298.00,
            invoice_description=f"{period} - The Kensington: Apartments.com Network 3 Gold Plus advertising",
            line_description=f"{period} - The Kensington - Network 3 Gold Plus advertising package",
            reference_number="285574361",
        )

    chadwell_lines = allocate_tax(
        [
            {"desc": 'Red Seal 4" jumbo wax bowl ring with sleeve', "base": 22.36, "qty": 4, "unit": 5.59, "gl": "6675", "taxable": True},
            {"desc": 'Fluidmaster stainless steel faucet supply line 3/8" comp x 1/2" IPS x 12"', "base": 19.77, "qty": 3, "unit": 6.59, "gl": "6675", "taxable": True},
            {"desc": "Tank-to-bowl closet bolts, 2-pack", "base": 8.78, "qty": 2, "unit": 4.39, "gl": "6675", "taxable": True},
            {"desc": "Bowl-to-floor toilet bolts, brass plated 5/16 x 2-1/4, 2-pack", "base": 5.07, "qty": 3, "unit": 1.69, "gl": "6675", "taxable": True},
            {"desc": "DAP Kwik Seal tub and tile white caulk, 5.5 oz", "base": 19.96, "qty": 4, "unit": 4.99, "gl": "6675", "taxable": True},
            {"desc": "9 watt LED A19 lamp, 4100K 800 lumens, 24-pack", "base": 71.89, "qty": 1, "unit": 71.89, "gl": "6627", "taxable": True},
        ],
        14.40,
    )
    inv_date = parse_date("02/26/2026")
    for i, line in enumerate(chadwell_lines, 1):
        rows.append(make_row(
            invoice_number="010890135",
            invoice_date=inv_date,
            due_date=add_month(inv_date),
            vendor=V_CHADWELL,
            invoice_description="02/26/26 - The Kensington: Chadwell maintenance supplies",
            line_item_number=i,
            property_abbreviation=TKA,
            gl_account=line["gl"],
            line_item_description=f"02/26/26 - The Kensington - {line['desc']}",
            amount=line["amount"],
            quantity=line["qty"],
            unit_price=line["unit"],
            reference_number="PO-3675",
            source_key="chadwell_010890135",
            document_url=links["chadwell_010890135"],
        ))

    handy_invoices = {
        "1029": {
            "date": "09/20/2025",
            "prop": TKA,
            "loc": "2102-9",
            "desc": "The Kensington Apt 2102-9: subfloor repair",
            "source": "handy_1029",
            "lines": [
                ("7572", "Subfloor repair labor and materials", 1920.00),
                ("7572", "Additional subfloor repair materials for unforeseen damage", 578.00),
            ],
        },
        "1063": {
            "date": "10/05/2025",
            "prop": TKA,
            "loc": "6B",
            "desc": "The Kensington Apt 6B: unit repaint",
            "source": "handy_1063",
            "lines": [("6760", "Unit repaint labor", 350.00)],
        },
        "1066": {
            "date": "10/05/2025",
            "prop": TKA,
            "loc": "2A",
            "desc": "The Kensington Apt 2A: final cleaning",
            "source": "handy_1066",
            "lines": [("6750", "Final clean after turn", 100.00)],
        },
        "1069": {
            "date": "10/05/2025",
            "prop": TLA,
            "loc": "6A",
            "desc": "Trinity Lofts Apt 6A: drywall repair",
            "source": "handy_1069",
            "lines": [("6780", "Drywall repair from water damage with trim replacement", 400.00)],
        },
        "1087": {
            "date": "10/16/2025",
            "prop": TKA,
            "loc": "",
            "desc": "The Kensington: mailbox installation",
            "source": "handy_1087",
            "lines": [("6615", "Mailbox installation with anchors and materials", 615.00)],
        },
        "1296": {
            "date": "02/14/2026",
            "prop": TKA,
            "loc": "5B",
            "desc": "The Kensington Apt 5B: repaint and final cleaning",
            "source": "handy_1296",
            "lines": [("6760", "Unit repaint labor", 300.00), ("6750", "Final clean after turn", 175.00)],
        },
        "1321": {
            "date": "03/02/2026",
            "prop": TKA,
            "loc": "203",
            "desc": "The Kensington Apt 203: repaint and final cleaning",
            "source": "handy_1321",
            "lines": [("6760", "Unit repaint labor", 350.00), ("6750", "Final clean after turn", 150.00)],
        },
        "1322": {
            "date": "03/02/2026",
            "prop": TKA,
            "loc": "203",
            "desc": "The Kensington Apt 203: repaint and final cleaning",
            "source": "handy_1322",
            "lines": [("6760", "Unit repaint labor", 300.00), ("6750", "Final clean after turn", 100.00)],
        },
        "1323": {
            "date": "03/02/2026",
            "prop": TKA,
            "loc": "203",
            "desc": "The Kensington Apt 203: repaint and final cleaning",
            "source": "handy_1323",
            "lines": [("6760", "Unit repaint labor", 300.00), ("6750", "Final clean after turn", 150.00)],
        },
    }
    for inv, info in handy_invoices.items():
        inv_date = parse_date(info["date"])
        for i, (gl, desc, amount) in enumerate(info["lines"], 1):
            rows.append(make_row(
                invoice_number=inv,
                invoice_date=inv_date,
                due_date=add_month(inv_date),
                vendor=V_HANDY,
                invoice_description=f"{inv_date.strftime('%m/%d/%y')} - {info['desc']}",
                line_item_number=i,
                property_abbreviation=info["prop"],
                location=info["loc"],
                gl_account=gl,
                line_item_description=f"{inv_date.strftime('%m/%d/%y')} - {info['desc']} - {desc}",
                amount=amount,
                quantity=1,
                unit_price=amount,
                source_key=info["source"],
                document_url=links[info["source"]],
            ))

    add_simple_invoice(
        rows,
        links=links,
        source_docs=source_docs,
        source_key="heavens_70242",
        invoice_number="70242",
        invoice_date=parse_date("09/03/2025"),
        due_date=parse_date("10/03/2025"),
        vendor=V_HEAVENS,
        prop=TKA,
        gl="6785",
        amount=450.00,
        invoice_description="09/03/25 - The Kensington: furniture and trash haul",
        line_description="09/03/25 - The Kensington - Furniture and trash haul from property",
    )

    cerv_invoices = [
        ("1237104", "08/26/2025", "09/25/2025", 70.00),
        ("1245963", "09/12/2025", "10/12/2025", 105.00),
        ("1261313", "10/12/2025", "11/11/2025", 105.00),
        ("1276179", "11/12/2025", "12/12/2025", 105.00),
        ("1290760", "12/12/2025", "01/11/2026", 105.00),
        ("1303747", "01/12/2026", "02/11/2026", 105.00),
    ]
    for inv, dt, due, amt in cerv_invoices:
        inv_date = parse_date(dt)
        add_simple_invoice(
            rows,
            links=links,
            source_docs=source_docs,
            source_key=f"cerv_{inv}",
            invoice_number=inv,
            invoice_date=inv_date,
            due_date=parse_date(due),
            vendor=V_CERV,
            prop=TKA,
            gl="6560",
            amount=amt,
            invoice_description=f"{month_label(inv_date)} - The Kensington: commercial premium pest protection",
            line_description=f"{month_label(inv_date)} - The Kensington - Commercial premium pest protection",
            reference_number="199873",
        )

    add_simple_invoice(
        rows,
        links=links,
        source_docs=source_docs,
        source_key="tenant_refund_temp54447402",
        invoice_number="TEMP54447402",
        invoice_date=parse_date("02/24/2026"),
        due_date=parse_date("03/26/2026"),
        vendor=V_TENANT_REFUND,
        prop=TKA,
        gl="2180",
        amount=500.00,
        invoice_description="02/24/26 - The Kensington Apt 2100-4: resident security deposit refund",
        line_description="02/24/26 - The Kensington Apt 2100-4 - Refund balance due from final account statement",
        location="2100-4",
    )

    add_mccoy_rows(rows, links)
    add_realpage_rows(rows, links)

    add_simple_invoice(
        rows,
        links=links,
        source_docs=source_docs,
        source_key="southern_19126",
        invoice_number="19126",
        invoice_date=parse_date("03/01/2026"),
        due_date=parse_date("03/01/2026"),
        vendor=V_SOUTHERN,
        prop=TKA,
        gl="6810",
        amount=354.60,
        invoice_description="Mar-26 - The Kensington: landscape maintenance agreement",
        line_description="Mar-26 - The Kensington - FY2025 landscape maintenance agreement",
        reference_number="4827980",
    )

    for inv, dt, due in [("L20959", "12/10/2025", "01/09/2026"), ("L22797", "01/10/2026", "02/09/2026"), ("L24470", "02/10/2026", "03/12/2026")]:
        add_simple_invoice(
            rows,
            links=links,
            source_docs=source_docs,
            source_key=f"verifast_{inv}",
            invoice_number=inv,
            invoice_date=parse_date(dt),
            due_date=parse_date(due),
            vendor=V_VERIFAST,
            prop=TKA,
            gl="6115",
            amount=28.75,
            invoice_description=f"{month_label(parse_date(dt))} - The Kensington: resident verification bundle",
            line_description=f"{month_label(parse_date(dt))} - The Kensington - Resident ID, income, employment, and bank verification bundle",
            quantity=23,
            unit_price=1.25,
        )

    zillow = [
        ("INV34935840", "12/31/2025", "01/30/2026", "12/15/2025", "12/31/2025", 38.39, "Base Package proration"),
        ("INV35306571", "01/31/2026", "03/02/2026", "01/01/2026", "01/31/2026", 70.00, "Base Package"),
        ("INV35702646", "02/28/2026", "03/30/2026", "02/01/2026", "02/28/2026", 70.00, "Base Package"),
    ]
    for inv, dt, due, start, end, amt, package in zillow:
        p = period_label(parse_date(start), parse_date(end))
        add_simple_invoice(
            rows,
            links=links,
            source_docs=source_docs,
            source_key=f"zillow_{inv}",
            invoice_number=inv,
            invoice_date=parse_date(dt),
            due_date=parse_date(due),
            vendor=V_ZILLOW,
            prop=TKA,
            gl="6335",
            amount=amt,
            invoice_description=f"{p} - The Kensington: Zillow Rentals - Zillow Rent Connect {package}",
            line_description=f"{p} - The Kensington - Zillow Rent Connect: {package}",
            reference_number="ZRN-7209535-ZRN-7204757",
        )

    return rows


def add_mccoy_rows(rows: list[dict[str, Any]], links: dict[str, str]) -> None:
    invoices = [
        ("mccoy_25427", "25427", "08/23/2025", "09/22/2025", TKA, [
            ("2100-2", "Candace Birdine", "Detainer warrant filing fee", 152.50),
            ("2100-2", "Candace Birdine", "Attorney fee - detainer warrant", 225.00),
            ("2100-2", "Candace Birdine", "After court letter fee", 50.00),
        ]),
        ("mccoy_27839", "27839", "09/18/2025", "10/18/2025", TKA, [
            ("2100-2", "Candace Birdine", "Physical eviction fee", 200.00),
            ("2100-2", "Candace Birdine", "Writ filing fee", 78.73),
            ("2100-2", "Candace Birdine", "Attorney fee - writ", 100.00),
            ("3A", "William Rice / Kristina Rice", "Detainer warrant filing fee", 152.50),
            ("3A", "William Rice / Kristina Rice", "Attorney fee - detainer warrant", 225.00),
        ]),
        ("mccoy_29074", "29074", "10/27/2025", "11/26/2025", TKA, [
            ("3A", "William Rice / Kristina Rice", "Nonsuit letter", 50.00),
        ]),
        ("mccoy_31739", "31739", "11/26/2025", "12/26/2025", TKA, [
            ("2102-5", "Shavonne Gray", "Detainer warrant filing fee", 152.50),
            ("2102-5", "Shavonne Gray", "Attorney fee - detainer warrant", 225.00),
            ("2102-9", "Shavonne Gray / Johnny R Valcourt", "Detainer warrant filing fee", 152.50),
            ("2102-9", "Shavonne Gray / Johnny R Valcourt", "Attorney fee - detainer warrant", 225.00),
        ]),
        ("mccoy_33588", "33588", "12/31/2025", "01/30/2026", TKA, [
            ("2102-5", "Shavonne Gray", "After court letter fee", 50.00),
            ("2102-9", "Shavonne Gray / Johnny R Valcourt", "After court letter fee", 50.00),
        ]),
        ("mccoy_35963", "35963", "02/28/2026", "03/30/2026", TKA, [
            ("2102-5", "Shavonne Gray", "Writ filing fee", 78.73),
            ("2102-5", "Shavonne Gray", "Attorney fee - writ", 100.00),
            ("2102-9", "Shavonne Gray / Johnny R Valcourt", "Writ filing fee", 78.73),
            ("2102-9", "Shavonne Gray / Johnny R Valcourt", "Attorney fee - writ", 100.00),
            ("2102-9", "Shavonne Gray / Johnny R Valcourt", "Physical eviction fee", 200.00),
            ("2102-7B", "Darren Michael Perry", "Detainer warrant filing fee", 160.86),
            ("2102-7B", "Darren Michael Perry", "Attorney fee - detainer warrant", 225.00),
            ("2102-7B", "Darren Michael Perry", "After court letter fee", 50.00),
        ]),
        ("mccoy_35757", "35757", "02/28/2026", "03/30/2026", TLA, [
            ("116", "Kenny Hardy", "Detainer warrant filing fee", 160.86),
            ("116", "Kenny Hardy", "Attorney fee - detainer warrant", 225.00),
            ("112", "Timothy Williams", "Detainer warrant filing fee", 160.86),
            ("112", "Timothy Williams", "Attorney fee - detainer warrant", 225.00),
            ("107", "Curtis Cathy", "Detainer warrant filing fee", 160.86),
            ("107", "Curtis Cathy", "Attorney fee - detainer warrant", 225.00),
            ("", "Jason Haase", "General letter", 50.00),
            ("222", "Isaiah Smith", "Detainer warrant filing fee", 160.86),
            ("222", "Isaiah Smith", "Attorney fee - detainer warrant", 225.00),
        ]),
        ("mccoy_39786", "39786", "03/18/2026", "04/17/2026", TLA, [
            ("112", "Timothy Williams", "After court letter fee", 50.00),
            ("107", "Curtis Cathy", "After court letter fee", 50.00),
            ("116", "Kenny Hardy", "After court letter fee", 50.00),
        ]),
        ("mccoy_40318", "40318", "04/20/2026", "05/20/2026", TLA, [
            ("222", "Isaiah Smith", "Bankruptcy motion for relief from stay - attorney fees", 600.00),
            ("222", "Isaiah Smith", "Bankruptcy motion for relief from stay - filing fee", 199.00),
        ]),
    ]
    for source_key, inv, dt, due, prop, line_defs in invoices:
        inv_date = parse_date(dt)
        prop_name = "Trinity Lofts" if prop == TLA else "The Kensington"
        summary = "; ".join(sorted({f"{loc} {tenant}".strip() for loc, tenant, _, _ in line_defs}))
        for i, (loc, tenant, desc, amount) in enumerate(line_defs, 1):
            rows.append(make_row(
                invoice_number=inv,
                invoice_date=inv_date,
                due_date=parse_date(due),
                vendor=V_MCCOY,
                invoice_description=f"{inv_date.strftime('%m/%d/%y')} - {prop_name}: legal eviction/court fees - {summary}",
                line_item_number=i,
                property_abbreviation=prop,
                location=loc,
                gl_account="6205",
                line_item_description=f"{inv_date.strftime('%m/%d/%y')} - {prop_name} {('Apt ' + loc) if loc else ''} - {tenant} - {desc}",
                amount=amount,
                quantity=1,
                unit_price=amount,
                source_key=source_key,
                document_url=links[source_key],
            ))


def add_realpage_rows(rows: list[dict[str, Any]], links: dict[str, str]) -> None:
    invoices = [
        ("I2504079315", "04/18/2025", "05/18/2025", "05/01/2025", "05/31/2025", 17.97, [
            ("6136", "Knock CRM monthly fees", 89.70, True),
            ("6136", "Property Management Essentials Plus monthly fees", 94.52, True),
            ("6136", "ODE Resident Services monthly fees", 6.00, False),
            ("6115", "LeasingDesk Screening Business Credit Report", 60.90, False),
            ("6115", "LeasingDesk Screening AI Enterprise", 113.40, False),
        ]),
        ("I2505131914", "05/16/2025", "06/15/2025", "06/01/2025", "06/30/2025", 17.97, [
            ("6136", "ODE Resident Services monthly fees", 6.00, False),
            ("6136", "Property Management Essentials Plus monthly fees", 94.52, True),
            ("6136", "Knock CRM monthly fees", 89.70, True),
            ("6115", "LeasingDesk Screening AI Enterprise", 272.16, False),
        ]),
        ("I2506139802", "06/18/2025", "07/18/2025", "07/01/2025", "07/31/2025", 17.97, [
            ("6136", "Property Management Essentials Plus monthly fees", 94.52, True),
            ("6136", "Knock CRM monthly fees", 89.70, True),
            ("6136", "ODE Resident Services monthly fees", 6.00, False),
            ("6115", "LeasingDesk Screening AI Enterprise", 68.04, False),
        ]),
        ("I2507083400", "07/17/2025", "08/17/2025", "08/01/2025", "08/31/2025", 17.97, [
            ("6136", "Knock CRM monthly fees", 89.70, True),
            ("6136", "Property Management Essentials Plus monthly fees", 94.52, True),
            ("6136", "ODE Resident Services monthly fees", 6.00, False),
        ]),
        ("I2508052387", "08/17/2025", "09/17/2025", "09/01/2025", "09/30/2025", 17.97, [
            ("6136", "ODE Resident Services monthly fees", 6.00, False),
            ("6136", "Property Management Essentials Plus monthly fees", 94.52, True),
            ("6136", "Knock CRM monthly fees", 89.70, True),
        ]),
        ("I2509133287", "09/17/2025", "10/17/2025", "10/01/2025", "10/31/2025", 17.97, [
            ("6136", "ODE Resident Services monthly fees", 6.00, False),
            ("6136", "Property Management Essentials Plus monthly fees", 94.52, True),
            ("6136", "Knock CRM monthly fees", 89.70, True),
        ]),
        ("I2510140691", "10/16/2025", "11/16/2025", "11/01/2025", "11/30/2025", 17.97, [
            ("6136", "Knock CRM monthly fees", 89.70, True),
            ("6136", "Property Management Essentials Plus monthly fees", 94.52, True),
            ("6136", "ODE Resident Services monthly fees", 6.00, False),
            ("6115", "LeasingDesk Screening AI Enterprise", 136.08, False),
        ]),
        ("I2511148820", "11/19/2025", "12/19/2025", "12/01/2025", "12/31/2025", 18.83, [
            ("6136", "Property Management Essentials Plus monthly fees", 99.01, True),
            ("6136", "Knock CRM monthly fees", 94.20, True),
            ("6136", "ODE Resident Services monthly fees", 6.30, False),
            ("6115", "LeasingDesk Screening AI Enterprise", 45.36, False),
        ]),
        ("I2512076391", "12/17/2025", "01/17/2026", "01/01/2026", "01/31/2026", 18.83, [
            ("6136", "ODE Resident Services monthly fees", 6.30, False),
            ("6136", "Knock CRM monthly fees", 94.20, True),
            ("6136", "Property Management Essentials Plus monthly fees", 99.01, True),
            ("6115", "LeasingDesk Screening Business Credit Report", 60.90, False),
        ]),
        ("I2601125816", "01/17/2026", "02/17/2026", "02/01/2026", "02/28/2026", 18.83, [
            ("6136", "Property Management Essentials Plus monthly fees", 99.01, True),
            ("6136", "Knock CRM monthly fees", 94.20, True),
            ("6136", "ODE Resident Services monthly fees", 6.30, False),
            ("6115", "LeasingDesk Screening AI Enterprise", 47.62, False),
        ]),
        ("I2602011770", "02/18/2026", "03/18/2026", "03/01/2026", "03/31/2026", 18.66, [
            ("6136", "Enterprise API Suite monthly fees for 02/02/26-02/28/26", 2.86, False),
            ("6136", "Enterprise API Suite monthly fees", 2.97, False),
            ("6136", "Knock CRM monthly fees", 93.30, True),
            ("6136", "ODE Resident Services monthly fees", 6.00, False),
            ("6136", "Property Management Essentials Plus monthly fees", 98.09, True),
            ("6115", "LeasingDesk Screening AI Enterprise", 47.18, False),
        ]),
    ]
    for inv, dt, due, start, end, tax, line_defs in invoices:
        inv_date = parse_date(dt)
        p = period_label(parse_date(start), parse_date(end))
        source_key = f"realpage_{inv}"
        line_dicts = allocate_tax(
            [{"gl": gl, "desc": desc, "base": base, "taxable": taxable} for gl, desc, base, taxable in line_defs],
            tax,
        )
        for i, line in enumerate(line_dicts, 1):
            rows.append(make_row(
                invoice_number=inv,
                invoice_date=inv_date,
                due_date=parse_date(due),
                vendor=V_REALPAGE,
                invoice_description=f"{p} - The Kensington: RealPage software and screening services",
                line_item_number=i,
                property_abbreviation=TKA,
                gl_account=line["gl"],
                line_item_description=f"{p} - The Kensington - RealPage {line['desc']}",
                amount=line["amount"],
                quantity=1,
                unit_price=line["amount"],
                reference_number="A2412000731",
                source_key=source_key,
                document_url=links[source_key],
            ))


def write_workbook(rows: list[dict[str, Any]]) -> None:
    shutil.copyfile(TEMPLATE_PATH, OUTPUT_PATH)
    wb = load_workbook(OUTPUT_PATH)
    ws = wb.active
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    col_by_header = {header: idx + 1 for idx, header in enumerate(headers)}
    for existing in range(ws.max_row, 1, -1):
        ws.delete_rows(existing)
    for r_idx, row in enumerate(rows, 2):
        for header in headers:
            if header not in col_by_header:
                continue
            value = row.get(header, "")
            cell = ws.cell(r_idx, col_by_header[header], value)
            if header in {"Amount", "Quantity", "Unit Price", "Tax"} and value not in ("", None):
                cell.number_format = '#,##0.00'
    ws.freeze_panes = "A2"
    wb.save(OUTPUT_PATH)


def write_audit(rows: list[dict[str, Any]], source_docs: dict[str, SourceDoc]) -> None:
    used_files = {source_docs[row["_source_key"]].filename for row in rows}
    all_files = {p.name for p in INVOICE_DIR.glob("*.pdf")}
    duplicate_or_support_only = sorted(all_files - used_files)
    totals: dict[str, float] = {}
    for row in rows:
        totals[row["Invoice Number"]] = totals.get(row["Invoice Number"], 0.0) + float(row["Amount"])
    with AUDIT_PATH.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["section", "key", "value"])
        writer.writerow(["summary", "unique_source_documents_used", len(used_files)])
        writer.writerow(["summary", "source_pdfs_in_folder", len(all_files)])
        writer.writerow(["summary", "template_rows", len(rows)])
        for inv, total in sorted(totals.items()):
            writer.writerow(["invoice_total", inv, f"{money(total):.2f}"])
        for filename in duplicate_or_support_only:
            writer.writerow(["duplicate_or_support_only_not_imported", filename, "Consolidated with matching invoice or not an AP invoice support row"])


def validate(rows: list[dict[str, Any]]) -> None:
    required = [
        "Invoice Number",
        "Bill or Credit",
        "Invoice Date",
        "Accounting Date",
        "Vendor",
        "Invoice Description",
        "Line Item Number",
        "Property Abbreviation",
        "GL Account",
        "Line Item Description",
        "Amount",
        "Expense Type",
        "Is Replacement Reserve",
        "Due Date",
        "Document Url",
    ]
    failures = []
    for idx, row in enumerate(rows, 2):
        for col in required:
            if row.get(col) in (None, ""):
                failures.append(f"row {idx}: missing {col}")
        if not str(row.get("Document Url", "")).startswith("https://"):
            failures.append(f"row {idx}: Document Url is not a Dropbox/http link")
        if not isinstance(row.get("Amount"), (int, float)) or math.isclose(float(row.get("Amount", 0)), 0.0):
            failures.append(f"row {idx}: invalid amount {row.get('Amount')!r}")
    if failures:
        raise AssertionError("\n".join(failures[:50]))


def main() -> None:
    source_docs = build_source_docs()
    links = upload_links(source_docs)
    rows = build_rows(links, source_docs)
    validate(rows)
    write_workbook(rows)
    write_audit(rows, source_docs)
    print(f"rows={len(rows)}")
    print(f"workbook={OUTPUT_PATH}")
    print(f"audit={AUDIT_PATH}")


if __name__ == "__main__":
    main()
