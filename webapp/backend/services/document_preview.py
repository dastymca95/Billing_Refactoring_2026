"""Convert uploaded files into JSON the frontend can render.

CSV / XLSX → table. PDF / image → served as raw bytes by a different
endpoint. Other types → metadata only.
"""

from __future__ import annotations

import csv
from io import StringIO
from pathlib import Path
from typing import Any

try:
    import openpyxl  # already in the project's deps
except ImportError:
    openpyxl = None  # type: ignore

try:
    from pypdf import PdfReader
except ImportError:
    PdfReader = None  # type: ignore


_MAX_PREVIEW_ROWS = 200
_MAX_PREVIEW_COLS = 50


def _open_csv_lenient(path: Path) -> StringIO:
    raw = path.read_bytes()
    for enc in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            return StringIO(raw.decode(enc))
        except UnicodeDecodeError:
            continue
    return StringIO(raw.decode("latin-1", errors="replace"))


def preview_csv(path: Path) -> dict[str, Any]:
    rows: list[list[str]] = []
    with _open_csv_lenient(path) as f:
        reader = csv.reader(f)
        for i, row in enumerate(reader):
            if i >= _MAX_PREVIEW_ROWS:
                break
            rows.append(row[:_MAX_PREVIEW_COLS])
    headers = rows[0] if rows else []
    data = rows[1:] if len(rows) > 1 else []
    return {
        "kind": "table",
        "filename": path.name,
        "headers": headers,
        "rows": data,
        "truncated_rows": _MAX_PREVIEW_ROWS,
    }


def preview_xlsx(path: Path) -> dict[str, Any]:
    if openpyxl is None:
        return {"kind": "metadata", "filename": path.name, "note": "openpyxl missing"}
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]
    headers: list[str] = []
    data: list[list[Any]] = []
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if r_idx == 1:
            headers = [("" if v is None else str(v)) for v in row][:_MAX_PREVIEW_COLS]
            continue
        if r_idx - 1 > _MAX_PREVIEW_ROWS:
            break
        data.append([("" if v is None else (v.isoformat() if hasattr(v, "isoformat") else v)) for v in row[:_MAX_PREVIEW_COLS]])
    return {
        "kind": "table",
        "filename": path.name,
        "sheet_name": sheet_name,
        "all_sheets": wb.sheetnames,
        "headers": headers,
        "rows": data,
        "truncated_rows": _MAX_PREVIEW_ROWS,
    }


def pdf_page_count(path: Path) -> int | None:
    """Return a PDF page count when the optional parser is available.

    This is metadata only for UI navigation. Failures are non-fatal so a
    damaged or unusually encoded PDF still uploads/renders through the normal
    preview path instead of blocking the operator.
    """
    if path.suffix.lower() != ".pdf" or PdfReader is None:
        return None
    try:
        reader = PdfReader(str(path))
        return max(1, len(reader.pages))
    except Exception:
        return None


def preview_metadata(path: Path) -> dict[str, Any]:
    """Fallback for binary files (PDF / image / docx). Frontend will use the
    /raw endpoint to actually render these."""
    out = {
        "kind": "binary",
        "filename": path.name,
        "extension": path.suffix.lower(),
        "size_bytes": path.stat().st_size,
        "note": "Preview is rendered by frontend via the /raw endpoint.",
    }
    page_count = pdf_page_count(path)
    if page_count is not None:
        out["page_count"] = page_count
    return out


def preview_file(path: Path) -> dict[str, Any]:
    suf = path.suffix.lower()
    if suf == ".csv":
        return preview_csv(path)
    if suf in (".xlsx", ".xls"):
        return preview_xlsx(path)
    return preview_metadata(path)
