"""Canonical invoice rules and universal reasoning helpers.

This layer turns the operator-authored Canonica rules workbook into runtime
rules for AI-assisted invoices. It deliberately sits after AI extraction and
before ResMan row construction: AI returns candidates, this module applies
business rules, internal references, and required-field validation.
"""

from __future__ import annotations

import copy
import re
from datetime import datetime, timedelta
from functools import lru_cache
from pathlib import Path
from typing import Any

import yaml

from .. import settings
from . import ai_mapping_review
from .description_builder import (
    build_contextual_one_off_line_description,
    build_one_off_content_summary,
)
from .service_invoice_gl_reasoning import build_gl_accounting_reasoning
from utils.text_normalization import (
    normalize_service_address_for_description,
    proper_case_preserve_acronyms,
)


CANONICAL_RULES_XLSX = settings.PROJECT_ROOT / "Canonica rules.xlsx"
CANONICAL_RULES_YAML = settings.PROJECT_ROOT / "config" / "canonical_rules.yaml"


class _NoAliasDumper(yaml.SafeDumper):
    def ignore_aliases(self, data: Any) -> bool:
        return True


COLUMN_KEY_TO_HEADER: dict[str, str] = {
    "invoice_number": "Invoice Number",
    "bill_or_credit": "Bill or Credit",
    "invoice_date": "Invoice Date",
    "accounting_date": "Accounting Date",
    "vendor": "Vendor",
    "invoice_description": "Invoice Description",
    "line_item_number": "Line Item Number",
    "property_abbreviation": "Property Abbreviation",
    "location": "Location",
    "gl_account": "GL Account",
    "line_item_description": "Line Item Description",
    "amount": "Amount",
    "expense_type": "Expense Type",
    "is_replacement_reserve": "Is Replacement Reserve",
    "payment_date": "Payment Date",
    "reference_number": "Reference Number",
    "payment_method": "Payment Method",
    "department": "Department",
    "due_date": "Due Date",
    "quantity": "Quantity",
    "unit_price": "Unit Price",
    "tax": "Tax",
    "received_date": "Received Date",
    "document_url": "Document Url",
}

HEADER_TO_COLUMN_KEY = {v.lower(): k for k, v in COLUMN_KEY_TO_HEADER.items()}

REQUIRED_TEMPLATE_COLUMNS: list[str] = [
    "Invoice Number",
    "Bill or Credit",
    "Invoice Date",
    "Accounting Date",
    "Vendor",
    "Invoice Description",
    "Line Item Number",
    "Property Abbreviation",
    "GL Account",
    "Line Item Description",
    "Amount",
    "Expense Type",
    "Is Replacement Reserve",
    "Due Date",
    "Document Url",
]

OPTIONAL_TEMPLATE_COLUMNS: list[str] = [
    "Location",
    "Payment Date",
    "Reference Number",
    "Payment Method",
    "Department",
    "Quantity",
    "Unit Price",
    "Tax",
    "Received Date",
]


def default_rules() -> dict[str, Any]:
    return {
        "version": 1,
        "source": {
            "excel_path": str(CANONICAL_RULES_XLSX),
            "runtime_source": str(CANONICAL_RULES_YAML),
            "notes": "Generated from Canonica rules.xlsx and completed with runtime defaults.",
        },
        "template_requirements": {
            "required_columns": REQUIRED_TEMPLATE_COLUMNS,
            "optional_columns": OPTIONAL_TEMPLATE_COLUMNS,
            "location_policy": "optional_valid_unit_only",
            "document_url_policy": "required_when_dropbox_available",
        },
        "template_columns": {
            "invoice_number": {
                "required": True,
                "source_priority": ["canonical_category_rule", "invoice_number_from_bill", "account_number_period_format"],
                "validation": {"cannot_be_blank": True},
            },
            "bill_or_credit": {
                "required": True,
                "default": "Bill",
                "credit_detection_keywords": ["credit memo", "credit", "refund"],
            },
            "invoice_date": {
                "required": True,
                "source_priority": ["explicit_invoice_date", "statement_date", "purchase_date", "ship_date"],
                "if_inferred": "flag_review",
            },
            "accounting_date": {"required": True, "default_rule": "same_as_invoice_date"},
            "vendor": {
                "required": True,
                "source": "vendor_list_exact_or_candidate_match",
                "if_unmatched": "vendor_mapping_required",
            },
            "invoice_description": {
                "required": True,
                "never_use_vague_ai_description": True,
                "formats": {
                    "utilities": "{service_period} - {service_address_or_property}",
                    "trash_collection_services": "{service_period} - {property_or_site_name}",
                    "pest_control": "{service_period} - {property_or_site_name}",
                    "landscaping": "{service_period} - {property_or_site_name}",
                    "marketing": "{invoice_date_short} - {vendor_name} - {main_item_or_category}",
                    "subscriptions": "{invoice_date_short} - {vendor_name} - {main_item_or_category}",
                    "other_infrequent": "{main_item_or_category}",
                    "unknown": "{invoice_date_short} - {vendor_name} - {main_item_or_category}",
                },
            },
            "line_item_number": {"required": True, "rule": "sequential_per_invoice"},
            "property_abbreviation": {
                "required": True,
                "source": "property_reference_match",
                "if_unmatched": "property_mapping_required",
            },
            "location": {
                "required": False,
                "source": "unit_info_valid_location_only",
                "never_use_raw_address": True,
            },
            "gl_account": {
                "required": True,
                "source": "chart_of_accounts_valid_numeric_only",
                "if_unmatched": "gl_mapping_required",
            },
            "line_item_description": {
                "required": True,
                "formats": {
                    "utilities": "{service_period} - {service_address_or_property} - {line_item_description}",
                    "trash_collection_services": "{service_period} - {property_or_site_name} - {line_item_description}",
                    "pest_control": "{service_period} - {property_or_site_name} - {line_item_description}",
                    "landscaping": "{service_period} - {property_or_site_name} - {line_item_description}",
                    "marketing": "{vendor_name} - {line_item_description}",
                    "subscriptions": "{vendor_name} - {line_item_description}",
                    "other_infrequent": "{line_item_description}",
                    "unknown": "{vendor_name} - {line_item_description}",
                },
            },
            "amount": {"required": True, "validation": {"numeric": True, "reconcile_to_invoice_total": True}},
            "expense_type": {"required": True, "default": "General"},
            "is_replacement_reserve": {"required": True, "default": False},
            "due_date": {"required": True, "source_priority": ["explicit_due_date", "autopay_date", "terms_plus_invoice_date"]},
            "document_url": {"required": True, "source": "dropbox_after_upload"},
        },
        "categories": {
            "utilities": {
                "labels": ["Utilities"],
                "vendor_keywords": ["utility", "utilities", "water", "electric", "wastewater", "fiber", "internet", "gas", "power"],
                "service_keywords": ["water", "sewer", "electric", "internet", "fiber", "gas", "wastewater"],
                "default_gl_candidates": {
                    "water": "6955",
                    "sewer": "6955",
                    "wastewater": "6955",
                    "internet": "6920",
                    "fiber": "6920",
                    "electric": "6950",
                    "gas": "6995",
                },
                "invoice_number_rule": "account_number_service_period",
                "location_policy": "valid_unit_if_present",
            },
            "pest_control": {
                "labels": ["Pest Control"],
                "vendor_keywords": ["pest", "termite", "exterminator"],
                "default_gl_candidates": {"default": "6760"},
                "location_policy": "property_level_blank_location_allowed",
            },
            "landscaping": {
                "labels": ["Landscaping"],
                "vendor_keywords": ["landscap", "lawn", "mulch", "tree"],
                "default_gl_candidates": {"default": "6810"},
                "location_policy": "property_level_blank_location_allowed",
            },
            "marketing": {
                "labels": ["Marketing"],
                "vendor_keywords": ["marketing", "advertising", "apartments.com", "costar"],
                "default_gl_candidates": {"default": "6630"},
                "location_policy": "property_level_blank_location_allowed",
            },
            "subscriptions": {
                "labels": ["Subscriptions", "Suscriptions"],
                "vendor_keywords": ["subscription", "software", "license", "saas"],
                "default_gl_candidates": {"default": "6665"},
                "location_policy": "property_level_blank_location_allowed",
            },
            "trash_collection_services": {
                "labels": ["Trash collections services"],
                "vendor_keywords": ["waste", "trash", "garbage", "refuse", "dumpster", "capital waste", "waste connections"],
                "service_keywords": ["trash", "waste", "container", "dumpster", "fuel recovery", "removal"],
                "default_gl_candidates": {"default": "6940", "fuel_recovery": "6940", "environmental": "6940"},
                "fee_handling": {"fuel_recovery_adjustment": "same_gl_as_service"},
                "ignore_line_keywords": ["previous balance", "balance forward", "payment received", "autopay", "amount enclosed"],
                "location_policy": "property_level_blank_location_allowed",
            },
            "other_infrequent": {
                "labels": ["Other infrequent / maintenance / supplies / purchases"],
                "vendor_keywords": ["supply", "supplies", "maintenance", "repair", "materials", "appliance", "parts"],
                "use_ai": True,
                "require_vendor_validation": True,
                "require_gl_validation": True,
                "location_policy": "valid_unit_if_present",
            },
            "unknown": {
                "labels": ["Unknown"],
                "vendor_keywords": [],
                "use_ai": True,
                "location_policy": "valid_unit_if_present",
            },
        },
    }


@lru_cache(maxsize=1)
def load_rules() -> dict[str, Any]:
    rules = default_rules()
    if CANONICAL_RULES_YAML.is_file():
        try:
            loaded = yaml.safe_load(CANONICAL_RULES_YAML.read_text(encoding="utf-8")) or {}
            if isinstance(loaded, dict):
                rules = _merge_dicts(rules, loaded)
        except Exception:
            return rules
    return rules


def reset_cache() -> None:
    load_rules.cache_clear()


def import_canonical_rules_from_excel(
    xlsx_path: Path | None = None,
    yaml_path: Path | None = None,
    *,
    write: bool = True,
) -> dict[str, Any]:
    """Import the operator workbook into editable runtime YAML."""
    from openpyxl import load_workbook

    xlsx_path = Path(xlsx_path or CANONICAL_RULES_XLSX)
    yaml_path = Path(yaml_path or CANONICAL_RULES_YAML)
    if not xlsx_path.is_file():
        raise FileNotFoundError(f"Canonical rules workbook not found: {xlsx_path}")

    rules = default_rules()
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb.active
    category_row = None
    for row_index, row in enumerate(ws.iter_rows(), start=1):
        values = [str(cell.value or "").strip() for cell in row]
        if any(_category_key(value) == "utilities" for value in values):
            category_row = row_index
            break
    if not category_row:
        wb.close()
        raise ValueError("Could not find category header row in Canonica rules.xlsx.")

    categories_by_col: dict[int, str] = {}
    for cell in ws[category_row]:
        label = str(cell.value or "").strip()
        key = _category_key(label)
        if key:
            categories_by_col[cell.column] = key
            rules["categories"].setdefault(key, {"labels": [label], "vendor_keywords": []})
            labels = rules["categories"][key].setdefault("labels", [])
            if label and label not in labels:
                labels.append(label)

    imported_rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=category_row + 1):
        label = str(row[1].value or "").strip() if len(row) >= 2 else ""
        if not label:
            continue
        col_key = _column_key(label)
        if not col_key:
            continue
        mandatory = "mandatory" in str(row[2].value or "").lower() if len(row) >= 3 else False
        col_rule = rules["template_columns"].setdefault(col_key, {})
        if mandatory:
            col_rule["required"] = True
        matrix: dict[str, str] = {}
        for column_index, cell in enumerate(row, start=1):
            cat_key = categories_by_col.get(column_index)
            if not cat_key:
                continue
            text = str(cell.value or "").strip()
            if text:
                matrix[cat_key] = text
        if matrix:
            col_rule["canonical_matrix"] = matrix
        imported_rows.append({
            "column": label,
            "key": col_key,
            "required": mandatory,
            "rules_by_category": matrix,
        })
    wb.close()

    rules["source"]["imported_rows"] = imported_rows
    rules["source"]["excel_path"] = str(xlsx_path)
    rules["source"]["runtime_source"] = str(yaml_path)

    if write:
        yaml_path.parent.mkdir(parents=True, exist_ok=True)
        yaml_path.write_text(
            yaml.dump(rules, Dumper=_NoAliasDumper, sort_keys=False, allow_unicode=False, width=120),
            encoding="utf-8",
        )
        reset_cache()
    return rules


def prompt_rules_summary() -> str:
    rules = load_rules()
    categories = rules.get("categories") or {}
    required = rules.get("template_requirements", {}).get("required_columns") or REQUIRED_TEMPLATE_COLUMNS
    category_lines = []
    for key, cfg in categories.items():
        labels = ", ".join(cfg.get("labels") or [key])
        keywords = ", ".join((cfg.get("vendor_keywords") or cfg.get("service_keywords") or [])[:8])
        category_lines.append(f"- {key}: labels={labels}; keywords={keywords}")
    return "\n".join(
        [
            "Canonical ResMan rules:",
            "AI returns extraction candidates only. The backend applies final ResMan formatting and validation.",
            "Required fields before export: " + ", ".join(required),
            "Never put a raw full address in Location. Location must be a valid unit/location or blank.",
            "GL Account must be a numeric code from the Chart of Accounts; vendor-side category text is source text only.",
            "Invoice Description and Line Item Description must follow the category-specific canonical formats.",
            "Classify economic nature from evidence: a metered recurring utility bill is not the same as a one-time plumbing, electrical, or sewer repair.",
            "A service address identifies where work occurred; it must never replace the content summary for a one-time invoice.",
            "Ignore previous balances, payments, autopay, amount enclosed, and remittance lines as expenses.",
            "Categories:",
            *category_lines,
        ]
    )


def canonicalize_normalized_invoice(
    normalized: dict[str, Any],
    *,
    references: dict[str, list[dict[str, Any]]] | None = None,
    rules_override: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Apply canonical rules to an already validated AI extraction."""
    out = copy.deepcopy(normalized)
    rules = copy.deepcopy(rules_override) if rules_override else load_rules()
    category = classify_invoice_category(out, rules=rules)
    out["category"] = category
    out["category_label"] = _category_label(category, rules)

    _apply_vendor_candidate(out)
    _apply_property_context(out, category, rules)
    _apply_category_invoice_number(out, category, rules)
    _apply_due_date_rules(out, category)
    _apply_gl_rules(out, category, rules)
    _apply_canonical_descriptions(out, category, rules)
    _apply_required_field_validation(out, rules)

    summary = dict(out.get("validation_summary") or {})
    summary["canonical_rules_applied"] = True
    summary["category"] = category
    summary["invoice_nature"] = out.get("invoice_nature") or "unknown"
    summary["invoice_nature_evidence"] = out.get("invoice_nature_evidence") or []
    summary["required_columns"] = rules.get("template_requirements", {}).get("required_columns", REQUIRED_TEMPLATE_COLUMNS)
    summary["blocking_required_fields"] = out.get("blocking_required_fields", [])
    out["validation_summary"] = summary
    return out


def classify_invoice_category(normalized: dict[str, Any], *, rules: dict[str, Any] | None = None) -> str:
    rules = rules or load_rules()
    nature, evidence = reason_invoice_nature(normalized)
    normalized["invoice_nature"] = nature
    normalized["invoice_nature_evidence"] = evidence
    explicit = _category_key(str(normalized.get("category") or ""))
    if nature == "one_time" and explicit == "utilities":
        return "other_infrequent"
    if nature == "utility_bill" and explicit in {"", "unknown", "other_infrequent"}:
        return "utilities"
    if explicit and explicit != "unknown" and explicit in rules.get("categories", {}):
        return explicit

    vendor_name = str(normalized.get("vendor_name") or normalized.get("raw_vendor_name") or "")
    vendor_rule_category = _vendor_yaml_category(vendor_name)
    if vendor_rule_category:
        mapped = _category_key(vendor_rule_category)
        if mapped in rules.get("categories", {}):
            if mapped == "utilities" and nature == "one_time":
                return "other_infrequent"
            return mapped

    haystack = " ".join(
        [
            vendor_name,
            str(normalized.get("invoice_description") or ""),
            str(normalized.get("service_address") or ""),
            " ".join(str(item.get("description") or "") for item in normalized.get("line_items") or [] if isinstance(item, dict)),
        ]
    ).lower()
    best_key = "unknown"
    best_score = 0
    for key, cfg in (rules.get("categories") or {}).items():
        score = 0
        for word in (cfg.get("vendor_keywords") or []) + (cfg.get("service_keywords") or []):
            word = str(word or "").lower().strip()
            if _category_keyword_matches(haystack, word):
                score += 2 if key == "trash_collection_services" and word in {"waste", "trash", "capital waste"} else 1
        if score > best_score:
            best_key, best_score = key, score
    if best_score:
        if best_key == "utilities" and nature == "one_time":
            return "other_infrequent"
        return best_key
    if nature == "one_time":
        return "other_infrequent"
    if nature == "utility_bill":
        return "utilities"
    return "unknown"


def reason_invoice_nature(normalized: dict[str, Any]) -> tuple[str, list[str]]:
    """Classify economic behavior from document evidence, independent of vendor rules."""
    vendor = ai_mapping_review.normalize_key(
        str(normalized.get("vendor_name") or normalized.get("raw_vendor_name") or "")
    )
    descriptions = " ".join(
        str(item.get("description") or item.get("source_line_description") or "")
        for item in normalized.get("line_items") or []
        if isinstance(item, dict)
    )
    content = ai_mapping_review.normalize_key(
        " ".join(
            [
                str(normalized.get("invoice_description") or ""),
                descriptions,
            ]
        )
    )
    combined = f"{vendor} {content}".strip()
    explicit = _category_key(str(normalized.get("category") or normalized.get("invoice_category") or ""))

    one_time_phrases = (
        "work order", "service call", "trip charge", "labor", "repair", "repaired",
        "replace", "replaced", "replacement", "install", "installed", "installation",
        "excavate", "excavated", "dug up", "cut out", "camera inspection", "diagnostic",
        "broken", "damaged", "leak", "clog", "pipe", "sewer line", "drain line",
        "restoration", "removal", "emergency service",
    )
    utility_structure_phrases = (
        "meter reading", "previous reading", "current reading", "usage charge", "amount used",
        "kwh", "kilowatt", "therm", "gallon", "billing cycle", "current charges",
        "total current charges", "balance forward", "previous balance", "account summary",
        "rate class", "days billed",
    )
    recurring_phrases = (
        "monthly", "recurring", "subscription", "scheduled service", "routine service",
        "contract period", "service agreement",
    )
    utility_provider_phrases = (
        "power system", "power company", "electric system", "electric company", "electric utility",
        "water authority", "water department", "water works", "waterworks", "wastewater department",
        "municipal utility", "municipal utilities", "utility district", "energy authority",
        "gas company", "gas and water", "gas & water", "telecommunications", "fiber optics",
    )

    action_hits = [phrase for phrase in one_time_phrases if _phrase_present(content, phrase)]
    utility_hits = [phrase for phrase in utility_structure_phrases if _phrase_present(combined, phrase)]
    recurring_hits = [phrase for phrase in recurring_phrases if _phrase_present(combined, phrase)]
    provider_hits = [phrase for phrase in utility_provider_phrases if _phrase_present(vendor, phrase)]
    contractor_identity = any(
        marker in vendor
        for marker in ("plumb", "rooter", "repair", "construction", "contractor", "landscap")
    )
    if (
        not provider_hits
        and not contractor_identity
        and re.search(r"\b(?:power|electric|water|gas|utility|utilities|wastewater)\b", vendor)
    ):
        provider_hits.append("utility-provider name")

    has_period = bool(
        normalized.get("service_period_start")
        or normalized.get("service_period_end")
        or normalized.get("service_period")
    )
    has_account = bool(str(normalized.get("account_number") or "").strip())
    if has_period:
        utility_hits.append("structured service period")
    if has_account:
        utility_hits.append("account number")

    evidence = [
        *(f"one_time:{value}" for value in action_hits[:5]),
        *(f"utility_structure:{value}" for value in utility_hits[:5]),
        *(f"utility_provider:{value}" for value in provider_hits[:3]),
        *(f"recurring:{value}" for value in recurring_hits[:3]),
    ]
    recurring_categories = {
        "pest_control", "landscaping", "marketing", "subscriptions",
        "trash_collection_services",
    }
    if explicit in recurring_categories or recurring_hits:
        return "recurring", evidence or [f"category:{explicit}"]
    if (provider_hits or explicit == "utilities") and len(utility_hits) >= 2:
        return "utility_bill", evidence
    if action_hits and len(utility_hits) < 2:
        return "one_time", evidence
    if provider_hits and utility_hits:
        return "utility_bill", evidence
    if len(action_hits) >= 2:
        return "one_time", evidence
    return "unknown", evidence


def _phrase_present(haystack: str, phrase: str) -> bool:
    if not haystack or not phrase:
        return False
    normalized_phrase = ai_mapping_review.normalize_key(phrase)
    return re.search(
        rf"(?<![a-z0-9]){re.escape(normalized_phrase)}(?![a-z0-9])",
        haystack,
    ) is not None


def required_columns() -> list[str]:
    return list(load_rules().get("template_requirements", {}).get("required_columns") or REQUIRED_TEMPLATE_COLUMNS)


def _apply_vendor_candidate(out: dict[str, Any]) -> None:
    vendor = str(out.get("vendor_name") or out.get("raw_vendor_name") or "").strip()
    if not vendor:
        return
    try:
        candidates = ai_mapping_review.vendor_candidates(vendor, limit=3)
    except Exception:
        return
    out["vendor_candidates"] = candidates.get("candidates") or []
    top = out["vendor_candidates"][0] if out["vendor_candidates"] else None
    if top and float(top.get("score") or 0) >= 0.90:
        out["vendor_name"] = top.get("vendor_name") or vendor


def _apply_property_context(out: dict[str, Any], category: str, rules: dict[str, Any]) -> None:
    if out.get("property_abbreviation"):
        if _category_allows_blank_location(category, rules) and _looks_like_raw_address(str(out.get("location") or "")):
            out["location"] = ""
        return
    query = str(out.get("property_candidate") or "")
    service_address = str(out.get("service_address") or "")
    try:
        candidates = ai_mapping_review.property_candidates(query=query, service_address=service_address, limit=5)
    except Exception:
        candidates = {"candidates": []}
    out["property_candidates"] = candidates.get("candidates") or []
    top = out["property_candidates"][0] if out["property_candidates"] else None
    if top and float(top.get("score") or 0) >= 0.78:
        out["property_abbreviation"] = top.get("property_abbreviation") or ""
        if _category_allows_blank_location(category, rules):
            out["location"] = ""
        else:
            out["location"] = top.get("location") or ""
        out["property_match"] = top
        _remove_issue(out, "property_mapping_required")
        if out.get("location"):
            _remove_issue(out, "location_unresolved")
    elif _category_allows_blank_location(category, rules):
        out["location"] = ""
    elif _looks_like_raw_address(str(out.get("location") or "")):
        out["location"] = ""


def _apply_category_invoice_number(out: dict[str, Any], category: str, rules: dict[str, Any]) -> None:
    source_invoice = str(out.get("source_invoice_number") or "").strip()
    if category != "utilities" and source_invoice:
        out["invoice_number"] = source_invoice
        out["invoice_number_policy_applied"] = False
        _remove_issue(out, "invoice_number_formatted_from_policy")
        _remove_issue(out, "invoice_number_missing")
        return
    current = str(out.get("invoice_number") or "").strip()
    if current and not str(out.get("invoice_number_generated") or "").lower() == "true":
        return
    cfg = (rules.get("categories") or {}).get(category) or {}
    if cfg.get("invoice_number_rule") == "account_number_service_period":
        account = str(out.get("account_number") or "").strip()
        month = _service_period_start_month3(out)
        year = _service_period_end_year2(out)
        if account and month and year:
            out["invoice_number"] = f"{account} {month} {year}"
            out["invoice_number_policy_applied"] = True
            _remove_issue(out, "invoice_number_missing")


def _apply_due_date_rules(out: dict[str, Any], category: str) -> None:
    if category == "utilities":
        return
    if category not in _NET30_CATEGORIES:
        return
    if str(out.get("due_date") or "").strip() and str(
        out.get("due_date_source") or ""
    ).strip() in {
        "explicit_due_date",
        "visible_payment_terms",
        "tenant_policy_from_due_date_text",
    }:
        return
    invoice_date = _parse_date(out.get("invoice_date"))
    if not invoice_date:
        return
    due_date = invoice_date + timedelta(days=30)
    out["due_date"] = due_date.strftime("%m/%d/%Y")
    _remove_issue(out, "required_due_date")
    _remove_issue(out, "due_date_missing")
    summary = dict(out.get("validation_summary") or {})
    summary["dates_valid"] = True
    summary["due_date_policy"] = "net_30_from_invoice_date"
    out["validation_summary"] = summary


def _apply_gl_rules(out: dict[str, Any], category: str, rules: dict[str, Any]) -> None:
    cfg = (rules.get("categories") or {}).get(category) or {}
    defaults = cfg.get("default_gl_candidates") or {}
    default_gl = str(defaults.get("default") or "").strip()
    for item in out.get("line_items") or []:
        if not isinstance(item, dict):
            continue
        reasoning = build_gl_accounting_reasoning(out, item, category)
        if reasoning:
            item["gl_accounting_reasoning"] = reasoning
            selected = str(reasoning.get("selected_gl_code") or "").strip()
            if selected:
                item["source_gl_candidate"] = item.get("gl_account_candidate")
                item["gl_account_candidate"] = selected
                item["gl_suggestion_source"] = "accounting_decision_engine"
        raw = str(item.get("gl_account_candidate") or "").strip()
        valid = ai_mapping_review.validate_gl_account(raw)
        if valid and ai_mapping_review.is_payable_gl_account(valid):
            item["gl_account_candidate"] = valid["gl_code"]
            item["gl_name"] = valid["gl_name"]
            continue

        desc_key = ai_mapping_review.normalize_key(str(item.get("description") or ""))
        matched_default = ""
        if category == "trash_collection_services":
            if any(k in desc_key for k in ("fuel recovery", "environmental", "trash", "waste", "container", "dumpster", "service")):
                matched_default = str(defaults.get("fuel_recovery") or default_gl)
        elif category == "utilities":
            for keyword, gl_code in defaults.items():
                if keyword != "default" and keyword in desc_key:
                    matched_default = str(gl_code)
                    break
        matched_default = matched_default or default_gl
        valid = ai_mapping_review.validate_gl_account(matched_default)
        if valid:
            item["gl_account_candidate"] = valid["gl_code"]
            item["gl_name"] = valid["gl_name"]
            item["gl_suggestion_source"] = "canonical_category_default"
            continue

        try:
            candidates = ai_mapping_review.gl_candidates(
                line_item_description=str(item.get("description") or ""),
                vendor_name=str(out.get("vendor_name") or ""),
                ai_suggested_gl=raw,
                limit=4,
            ).get("candidates") or []
        except Exception:
            candidates = []
        item["gl_candidates"] = candidates
        top = candidates[0] if candidates else None
        top_account = ai_mapping_review.validate_gl_account(top.get("gl_code") or top.get("gl_account") or "") if top else None
        if (
            top
            and top.get("valid")
            and float(top.get("score") or 0) >= 0.88
            and top_account
            and ai_mapping_review.is_payable_gl_account(top_account)
        ):
            item["gl_account_candidate"] = top_account["gl_code"]
            item["gl_name"] = top_account["gl_name"]
            item["gl_suggestion_source"] = "canonical_candidate_engine"


def _reasoning_candidates(reasoning: dict[str, Any]) -> list[dict[str, Any]]:
    candidates: list[dict[str, Any]] = []
    selected = str(reasoning.get("selected_gl_code") or "")
    for source in ("alternatives", "rejected_alternatives"):
        for item in reasoning.get(source) or []:
            if not isinstance(item, dict):
                continue
            code = str(item.get("gl_code") or "").strip()
            if not code or code == selected:
                continue
            account = ai_mapping_review.validate_gl_account(code)
            if not account:
                continue
            candidates.append({
                "gl_account": account["gl_code"],
                "gl_code": account["gl_code"],
                "gl_name": account["gl_name"],
                "score": 0.65 if source == "alternatives" else 0.30,
                "reason": item.get("reason") or "",
                "valid": True,
            })
    seen: set[str] = set()
    out: list[dict[str, Any]] = []
    for candidate in candidates:
        code = str(candidate.get("gl_code") or "")
        if code in seen:
            continue
        seen.add(code)
        out.append(candidate)
    return out[:4]


def _semantic_expense_gl(
    normalized: dict[str, Any],
    item: dict[str, Any],
    category: str,
) -> tuple[str, str] | None:
    """Choose a GL from the purchased work, not incidental words in its narrative."""
    if category != "other_infrequent" or normalized.get("invoice_nature") != "one_time":
        return None
    text = ai_mapping_review.normalize_key(
        " ".join(
            [
                str(normalized.get("vendor_name") or normalized.get("raw_vendor_name") or ""),
                str(normalized.get("invoice_description") or ""),
                str(item.get("description") or item.get("source_line_description") or ""),
                str(normalized.get("_document_text") or ""),
            ]
        )
    )
    work_markers = (
        "labor", "repair", "repaired", "replace", "replaced", "install", "installed",
        "service call", "excavate", "excavated", "dug up", "cut out", "diagnostic",
    )
    is_work = any(_phrase_present(text, marker) for marker in work_markers)
    if not is_work:
        return None

    plumbing = any(
        _phrase_present(text, marker)
        for marker in ("plumb", "sewer line", "drain line", "water line", "pipe", "rooter")
    )
    if plumbing:
        major_replacement = any(
            _phrase_present(text, marker)
            for marker in ("complete replacement", "full replacement", "entire line replacement", "replace entire")
        )
        if major_replacement:
            return "7560", "Major plumbing-system replacement is the purchased work."
        return "6565", "Plumbing contract/repair is the purchased work; incidental cause terms are ignored."

    if any(_phrase_present(text, marker) for marker in ("electrician", "electrical repair", "electrical work")):
        return "6540", "Electrical contract work is the purchased service."
    hvac_markers = (
        "hvac", "air conditioner", "a c", "heat pump", "furnace", "air handler",
        "condensing unit", "refrigerant", "freon", "thermostat",
    )
    if any(_phrase_present(text, marker) for marker in hvac_markers):
        full_replacement = any(
            _phrase_present(text, marker)
            for marker in (
                "complete hvac system replacement",
                "complete replacement",
                "full replacement",
                "replace entire",
                "remove old air handler and replace",
                "remove old a c condensing unit and replace",
            )
        ) or all(
            _phrase_present(text, marker)
            for marker in ("condensing unit", "disconnect", "thermostat")
        )
        if full_replacement:
            return "7544", "The purchased work is a complete HVAC equipment replacement."
        return "6555", "HVAC contract work is the purchased service."
    if any(_phrase_present(text, marker) for marker in ("appliance repair", "refrigerator repair", "washer repair", "dryer repair")):
        return "6505", "Appliance contract work is the purchased service."
    return None


def _apply_canonical_descriptions(out: dict[str, Any], category: str, rules: dict[str, Any]) -> None:
    invoice_format = (
        rules.get("template_columns", {})
        .get("invoice_description", {})
        .get("formats", {})
        .get(category)
    ) or rules.get("template_columns", {}).get("invoice_description", {}).get("formats", {}).get("unknown")
    line_format = (
        rules.get("template_columns", {})
        .get("line_item_description", {})
        .get("formats", {})
        .get(category)
    ) or rules.get("template_columns", {}).get("line_item_description", {}).get("formats", {}).get("unknown")
    if _uses_monthly_service_description(out, category):
        invoice_format = "{service_period} - {main_item_or_category}"
        line_format = "{service_period} - {line_item_description}"

    context = _format_context(out)
    sample_item = next((item for item in out.get("line_items") or [] if isinstance(item, dict)), {})
    invoice_description = _render_format(invoice_format, out, sample_item, context)
    if invoice_description:
        out["canonical_invoice_description"] = invoice_description[:180]

    for item in out.get("line_items") or []:
        if not isinstance(item, dict):
            continue
        rendered = _render_format(line_format, out, item, context)
        if rendered:
            item["canonical_line_item_description"] = rendered[:240]


def _apply_required_field_validation(out: dict[str, Any], rules: dict[str, Any]) -> None:
    blocking: list[str] = []
    required = rules.get("template_requirements", {}).get("required_columns") or REQUIRED_TEMPLATE_COLUMNS
    field_values = {
        "Invoice Number": out.get("invoice_number"),
        "Bill or Credit": out.get("bill_or_credit") or "Bill",
        "Invoice Date": out.get("invoice_date"),
        "Accounting Date": out.get("invoice_date"),
        "Vendor": out.get("vendor_name"),
        "Invoice Description": out.get("canonical_invoice_description") or out.get("invoice_description"),
        "Due Date": out.get("due_date"),
    }
    for header, value in field_values.items():
        if header in required and not _clean(value):
            blocking.append(header)

    items = [item for item in out.get("line_items") or [] if isinstance(item, dict)]
    if not items and "Line Item Description" in required:
        blocking.append("Line Items")
    for item in items:
        row_values = {
            "Property Abbreviation": out.get("property_abbreviation"),
            "GL Account": item.get("gl_account_candidate"),
            "Line Item Description": item.get("canonical_line_item_description") or item.get("description"),
            "Amount": item.get("amount"),
            "Expense Type": item.get("expense_type") or "General",
            "Is Replacement Reserve": False if item.get("is_replacement_reserve") is None else item.get("is_replacement_reserve"),
            "Line Item Number": 1,
        }
        for header, value in row_values.items():
            if header in required and value in ("", None):
                blocking.append(header)

    seen: set[str] = set()
    out["blocking_required_fields"] = [x for x in blocking if not (x in seen or seen.add(x))]
    summary = dict(out.get("validation_summary") or {})
    reconciliation_blocked = summary.get("total_reconciliation_passed") is False
    export_blocked = bool(out["blocking_required_fields"]) or reconciliation_blocked
    summary["required_fields_present"] = not export_blocked
    summary["valid"] = bool(summary.get("valid", True)) and not export_blocked
    summary["export_blocked"] = export_blocked
    out["validation_summary"] = summary
    for header in out["blocking_required_fields"]:
        code = _required_code(header)
        _add_issue(out, code, f"{header} is required by Canonical Rules before this invoice can be exported.", "high")


def _render_format(fmt: str, normalized: dict[str, Any], item: dict[str, Any], context: dict[str, str]) -> str:
    if not fmt:
        return ""
    source_line = _line_item_description_for_format(normalized, item)
    values = {
        **context,
        "vendor_name": _display_vendor(normalized),
        "invoice_date_short": _short_date(normalized.get("invoice_date")),
        "line_item_description": source_line or "Invoice total",
        "line_item_description_short": _concise_item(source_line),
        "main_item_or_category": _main_item_or_category(normalized, source_line),
        "amount": _format_amount(item.get("amount")),
    }
    rendered = fmt
    for key, value in values.items():
        rendered = rendered.replace("{" + key + "}", value)
    rendered = re.sub(r"\s+", " ", rendered).strip()
    rendered = re.sub(r"\s+-\s+-\s+", " - ", rendered)
    rendered = rendered.strip(" -")
    return proper_case_preserve_acronyms(rendered)


def _line_item_description_for_format(normalized: dict[str, Any], item: dict[str, Any]) -> str:
    verified_normalized = _clean(item.get("normalized_source_description"))
    if item.get("row_identity_evidence") and verified_normalized:
        # The raw source description remains immutable. When a bounded
        # pixel-first verifier corrected only the handwritten row identity,
        # canonical display text must consume that normalized fact rather
        # than reintroduce the superseded OCR label.
        return verified_normalized
    category = str(normalized.get("category") or "").strip().lower()
    if category in {"other_infrequent", "unknown"} and not _uses_monthly_service_description(
        normalized,
        category,
    ):
        text = build_contextual_one_off_line_description(normalized, item)
    else:
        text = _clean(item.get("description"))
    if _uses_monthly_service_description(normalized, str(normalized.get("category") or "")):
        text = _strip_trailing_source_date(text)
    return text


def _main_item_or_category(normalized: dict[str, Any], source_line: str) -> str:
    category = str(normalized.get("category") or "").strip().lower()
    if category in {"other_infrequent", "unknown"} and not _uses_monthly_service_description(
        normalized,
        category,
    ):
        summary = build_one_off_content_summary(normalized)
        if summary:
            return summary
    summary = _clean(normalized.get("invoice_description"))
    summary_norm = ai_mapping_review.normalize_key(summary)
    generic = {
        "",
        "invoice",
        "invoice total",
        "general invoice",
        "miscellaneous",
        "maintenance supplies",
        "hardware and miscellaneous items",
    }
    if summary_norm not in generic:
        return _concise_item(summary)
    return _concise_item(source_line) or _clean(normalized.get("category_label"))


def _strip_trailing_source_date(value: str) -> str:
    text = _clean(value)
    if not text:
        return ""
    text = re.sub(
        r"\s*(?:[-–—]\s*)?\b\d{1,2}[/-]\d{1,2}(?:[/-]\d{2,4})?\b\s*$",
        "",
        text,
    )
    return _clean(text)


def _uses_monthly_service_description(normalized: dict[str, Any], category: str) -> bool:
    if category in _ALWAYS_MONTHLY_DESCRIPTION_CATEGORIES:
        return True
    if category != "other_infrequent":
        return False
    haystack = ai_mapping_review.normalize_key(
        " ".join(
            [
                str(normalized.get("vendor_name") or ""),
                str(normalized.get("raw_vendor_name") or ""),
                str(normalized.get("invoice_description") or ""),
                " ".join(
                    str(item.get("description") or "")
                    for item in normalized.get("line_items") or []
                    if isinstance(item, dict)
                ),
            ]
        )
    )
    return any(
        token in haystack
        for token in (
            "monthly",
            "recurring",
            "subscription",
            "scheduled service",
            "routine service",
            "service period",
            "contract period",
        )
    )


def _format_context(normalized: dict[str, Any]) -> dict[str, str]:
    category = str(normalized.get("category") or "").strip()
    period = _service_period_label(normalized, category)
    prop_or_site = _property_or_site(normalized)
    service_address = normalize_service_address_for_description(normalized.get("service_address"))
    service_address_or_property = service_address or prop_or_site
    return {
        "service_period": period,
        "property_or_site_name": prop_or_site,
        "service_address_or_property": service_address_or_property,
        "property_abbreviation": _clean(normalized.get("property_abbreviation")),
        "service_address": service_address,
    }


def _property_or_site(normalized: dict[str, Any]) -> str:
    candidate = _clean(normalized.get("property_candidate"))
    if candidate and candidate.lower() != _clean(normalized.get("property_abbreviation")).lower():
        return candidate
    match = normalized.get("property_match") or {}
    for key in ("property_name", "Property Name"):
        value = _clean(match.get(key)) if isinstance(match, dict) else ""
        if value:
            return value
    address = _clean(normalized.get("service_address"))
    if address:
        return _first_address_phrase(address)
    return _clean(normalized.get("property_abbreviation"))


_ALWAYS_MONTHLY_DESCRIPTION_CATEGORIES = {
    "pest_control",
    "landscaping",
}

_NET30_CATEGORIES = {
    "pest_control",
    "landscaping",
    "other_infrequent",
    "unknown",
}


def _service_period_label(normalized: dict[str, Any], category: str = "") -> str:
    if _uses_monthly_service_description(normalized, category):
        monthly = _monthly_period_label(normalized)
        if monthly:
            return monthly
    start = _short_date(normalized.get("service_period_start"))
    end = _short_date(normalized.get("service_period_end"))
    if start and end and start == end:
        return start
    return f"{start}-{end}" if start and end else ""


def _category_keyword_matches(haystack: str, keyword: str) -> bool:
    needle = ai_mapping_review.normalize_key(keyword)
    if not needle:
        return False
    hay = ai_mapping_review.normalize_key(haystack)
    if not hay:
        return False
    if needle in {"landscap", "util"}:
        return needle in hay
    return re.search(rf"(?<![a-z0-9]){re.escape(needle)}(?![a-z0-9])", hay) is not None


def _monthly_period_label(normalized: dict[str, Any]) -> str:
    parsed = (
        _parse_date(normalized.get("invoice_date"))
        or _parse_date(normalized.get("service_period_end"))
        or _parse_date(normalized.get("service_period_start"))
    )
    return parsed.strftime("%b-%y") if parsed else ""


def _service_period_start_month3(normalized: dict[str, Any]) -> str:
    value = _parse_date(normalized.get("service_period_start") or normalized.get("invoice_date"))
    return value.strftime("%b") if value else ""


def _service_period_end_year2(normalized: dict[str, Any]) -> str:
    value = _parse_date(normalized.get("service_period_end") or normalized.get("invoice_date"))
    return value.strftime("%y") if value else ""


def _short_date(value: Any) -> str:
    parsed = _parse_date(value)
    return parsed.strftime("%m/%d/%y") if parsed else _clean(value)


def _parse_date(value: Any) -> datetime | None:
    text = _clean(value)
    if not text:
        return None
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%d-%b-%Y", "%d-%b-%y"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def _format_amount(value: Any) -> str:
    try:
        return f"{float(value or 0):.2f}"
    except (TypeError, ValueError):
        return ""


def _display_vendor(normalized: dict[str, Any]) -> str:
    return _clean(normalized.get("vendor_name") or normalized.get("raw_vendor_name"))


def _concise_item(value: str) -> str:
    text = re.sub(r"\s+", " ", value or "").strip()
    if not text:
        return ""
    generic = {"invoice total", "miscellaneous", "maintenance supplies", "hardware and miscellaneous items"}
    if text.lower() in generic:
        return ""
    words = text.split()
    return " ".join(words[:8])[:72]


def _first_address_phrase(value: str) -> str:
    return value.split(",", 1)[0].strip()


def _category_allows_blank_location(category: str, rules: dict[str, Any]) -> bool:
    policy = ((rules.get("categories") or {}).get(category) or {}).get("location_policy")
    return policy == "property_level_blank_location_allowed"


def _looks_like_raw_address(value: str) -> bool:
    text = value.strip()
    return bool(re.search(r"\d+\s+\w+", text) and ("," in text or re.search(r"\b(st|street|ave|road|rd|dr|ct|ln)\b", text, re.I)))


def _required_code(header: str) -> str:
    return "required_" + re.sub(r"[^a-z0-9]+", "_", header.lower()).strip("_")


def _add_issue(out: dict[str, Any], code: str, message: str, severity: str = "medium") -> None:
    issues = list(out.get("manual_review_issues") or [])
    if not any(issue.get("code") == code for issue in issues if isinstance(issue, dict)):
        issues.append({"code": code, "message": message, "severity": severity})
    out["manual_review_issues"] = issues
    codes = list(out.get("manual_review_codes") or [])
    if code not in codes:
        codes.append(code)
    out["manual_review_codes"] = codes
    reasons = list(out.get("manual_review_reasons") or [])
    if message not in reasons:
        reasons.append(message)
    out["manual_review_reasons"] = reasons


def _remove_issue(out: dict[str, Any], code: str) -> None:
    removed_messages = {
        i.get("message")
        for i in out.get("manual_review_issues") or []
        if isinstance(i, dict) and i.get("code") == code and i.get("message")
    }
    out["manual_review_issues"] = [i for i in out.get("manual_review_issues") or [] if not isinstance(i, dict) or i.get("code") != code]
    out["manual_review_codes"] = [c for c in out.get("manual_review_codes") or [] if c != code]
    out["manual_review_reasons"] = [r for r in out.get("manual_review_reasons") or [] if r not in removed_messages]


def _clean(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _column_key(label: str) -> str:
    cleaned = _clean(label).lower()
    direct = HEADER_TO_COLUMN_KEY.get(cleaned)
    if direct:
        return direct
    return re.sub(r"[^a-z0-9]+", "_", cleaned).strip("_")


def _category_key(label: str) -> str:
    text = ai_mapping_review.normalize_key(label)
    if not text:
        return ""
    if "util" in text or "internet telecom" in text:
        return "utilities"
    if "pest" in text:
        return "pest_control"
    if "landscap" in text:
        return "landscaping"
    if "marketing" in text:
        return "marketing"
    if "subscription" in text or "suscription" in text:
        return "subscriptions"
    if "trash" in text or "waste" in text or "garbage" in text:
        return "trash_collection_services"
    if any(
        token in text
        for token in (
            "maintenance",
            "supplies",
            "purchase",
            "unfrequent",
            "infrequent",
            "repair",
            "remodel",
            "construction",
            "restoration",
            "cleaning",
            "appliance",
        )
    ):
        return "other_infrequent"
    if text == "unknown":
        return "unknown"
    return re.sub(r"[^a-z0-9]+", "_", text).strip("_")


def _category_label(category: str, rules: dict[str, Any]) -> str:
    labels = ((rules.get("categories") or {}).get(category) or {}).get("labels") or []
    return str(labels[0]) if labels else category.replace("_", " ").title()


def _vendor_yaml_category(vendor_name: str) -> str:
    key = ai_mapping_review.mapping_key(vendor_name)
    path = settings.VENDORS_DIR / f"{key}.yaml"
    if not path.is_file():
        return ""
    try:
        data = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    except Exception:
        return ""
    identity = data.get("vendor_identity") or {}
    return str(identity.get("category") or "")


def _merge_dicts(base: dict[str, Any], override: dict[str, Any]) -> dict[str, Any]:
    merged = copy.deepcopy(base)
    for key, value in override.items():
        if isinstance(value, dict) and isinstance(merged.get(key), dict):
            merged[key] = _merge_dicts(merged[key], value)
        else:
            merged[key] = copy.deepcopy(value)
    return merged


__all__ = [
    "CANONICAL_RULES_XLSX",
    "CANONICAL_RULES_YAML",
    "REQUIRED_TEMPLATE_COLUMNS",
    "canonicalize_normalized_invoice",
    "classify_invoice_category",
    "default_rules",
    "import_canonical_rules_from_excel",
    "load_rules",
    "prompt_rules_summary",
    "reason_invoice_nature",
    "required_columns",
    "reset_cache",
]
