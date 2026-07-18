"""Universal file ingestion normalization.

This service stops invoice processing from treating every file type as a
special case. It extracts file-level text/table/page candidates and returns a
single DocumentCandidate shape that downstream reasoning can consume. It does
not choose vendors, GL accounts, properties, or ResMan rows.
"""

from __future__ import annotations

import csv
import mimetypes
import re
import zipfile
from dataclasses import asdict, dataclass, field
from io import StringIO
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

from .. import settings


PDF_TEXT_MIN_CHARS = 80
MAX_TABLE_ROWS = 500
MAX_TABLE_COLS = 80
MAX_EXCEL_CELLS = 25_000

IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp", ".tif", ".tiff"}
EXCEL_EXTENSIONS = {".xlsx", ".xls"}
WORD_EXTENSIONS = {".docx", ".doc"}


@dataclass
class TextBlockCandidate:
    text: str
    bbox: dict[str, float] | None = None
    confidence: float | None = None
    source: str = "pdf_text"


@dataclass
class PageCandidate:
    page_number: int
    text: str = ""
    text_quality_score: float = 0.0
    ocr_confidence: float | None = None
    image_ref: str | None = None
    image_path: str | None = None
    temp_image_path: str | None = None
    width: float | None = None
    height: float | None = None
    blocks: list[TextBlockCandidate | dict[str, Any]] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


@dataclass
class TableCandidate:
    source: str
    source_file: str = ""
    source_type: str = ""
    sheet_name: str | None = None
    page_number: int | None = None
    table_index: int = 1
    headers: list[str] = field(default_factory=list)
    rows: list[list[Any]] = field(default_factory=list)
    columns: list[str] = field(default_factory=list)
    confidence: float | None = None
    raw_cells: list[list[Any]] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


@dataclass
class ImageCandidate:
    source_file: str
    page_number: int | None = None
    image_path: str | None = None
    width: float | None = None
    height: float | None = None
    purpose: str = "original_image"
    warnings: list[str] = field(default_factory=list)


@dataclass
class DocumentCandidate:
    source_file: str
    source_type: str
    source_path: str = ""
    mime_type: str = ""
    file_size_bytes: int = 0
    page_count: int = 0
    sheet_count: int = 0
    vendor_hint: str = ""
    category_hint: str = ""
    document_text: str = ""
    text_quality_score: float = 0.0
    needs_ocr: bool = False
    needs_vision: bool = False
    pages: list[PageCandidate] = field(default_factory=list)
    tables: list[TableCandidate] = field(default_factory=list)
    images: list[ImageCandidate | dict[str, Any]] = field(default_factory=list)
    metadata: dict[str, Any] = field(default_factory=dict)
    extraction_quality: dict[str, Any] = field(default_factory=dict)
    warnings: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


def ingest_document(
    path: Path,
    *,
    vendor_hint: str = "",
    allow_ocr: bool = True,
    allow_vision_hint: bool = True,
    max_pages: int | None = None,
) -> DocumentCandidate:
    """Normalize an uploaded file into a DocumentCandidate."""
    path = Path(path)
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        candidate = _ingest_pdf(
            path,
            vendor_hint=vendor_hint,
            max_pages=max_pages,
            allow_ocr=allow_ocr,
        )
    elif suffix in IMAGE_EXTENSIONS:
        candidate = _ingest_image(path, vendor_hint=vendor_hint, allow_ocr=allow_ocr)
    elif suffix in EXCEL_EXTENSIONS:
        candidate = _ingest_excel(path, vendor_hint=vendor_hint)
    elif suffix in WORD_EXTENSIONS:
        candidate = _ingest_word(path, vendor_hint=vendor_hint)
    elif suffix == ".csv":
        candidate = _ingest_csv(path, vendor_hint=vendor_hint)
    else:
        candidate = _ingest_unknown(path, vendor_hint=vendor_hint)
    return _finalize_candidate(candidate, path, allow_vision_hint=allow_vision_hint)


def detect_source_type(path: Path) -> str:
    """Return the fast extension/content-based source type without parsing fully."""
    path = Path(path)
    if _is_internal_template(path):
        return "internal_template"
    if path.suffix.lower() == ".pdf":
        return _quick_pdf_source_type(path)
    return _source_type_from_extension(path)


def detect_file_support(path: Path) -> dict[str, Any]:
    """Fast UI metadata for upload/file lists.

    This intentionally avoids OCR and workbook parsing. It is safe to call from
    the batch-list endpoint, which is latency-sensitive.
    """
    suffix = Path(path).suffix.lower()
    if _is_internal_template(Path(path)):
        return {
            "source_type": "internal_template",
            "file_support_status": "unsupported",
            "file_support_label": _source_type_label("internal_template"),
            "file_support_reason": "Internal ResMan template files are ignored as invoice sources.",
        }
    source_type = _source_type_from_extension(path)
    status = "supported" if source_type != "unknown" else "unsupported"
    reason = "Supported by universal ingestion."
    if suffix == ".pdf":
        source_type = _quick_pdf_source_type(path)
        reason = "PDF has extractable text." if source_type == "pdf_digital" else "PDF may need OCR or vision."
        if source_type == "pdf_scanned":
            status = "limited"
            reason = "Scanned or weak-text PDF; OCR or vision may be recommended."
    elif source_type in {"image", "screenshot"}:
        status = "limited"
        reason = "Image source; OCR or vision may be recommended."
    elif suffix == ".xls":
        status = "limited"
        reason = "Legacy .xls may be unsupported unless optional readers are installed."
    elif suffix == ".doc":
        source_type = "unknown"
        reason = "Legacy .doc is not supported; use .docx."
        status = "unsupported"
    elif source_type == "unknown":
        reason = "This file type is not supported by universal ingestion."
    return {
        "source_type": source_type,
        "file_support_status": status,
        "file_support_label": _source_type_label(source_type),
        "file_support_reason": reason,
    }


def document_candidate_from_dict(value: dict[str, Any]) -> DocumentCandidate:
    """Rehydrate a candidate when a caller has JSON/dict state."""
    pages = [PageCandidate(**p) for p in value.get("pages") or []]
    tables = [TableCandidate(**t) for t in value.get("tables") or []]
    images = [
        ImageCandidate(**img) if isinstance(img, dict) and "source_file" in img else img
        for img in value.get("images") or []
    ]
    return DocumentCandidate(
        source_file=str(value.get("source_file") or ""),
        source_type=str(value.get("source_type") or "unknown"),
        source_path=str(value.get("source_path") or ""),
        mime_type=str(value.get("mime_type") or ""),
        file_size_bytes=int(value.get("file_size_bytes") or 0),
        page_count=int(value.get("page_count") or 0),
        sheet_count=int(value.get("sheet_count") or 0),
        vendor_hint=str(value.get("vendor_hint") or ""),
        category_hint=str(value.get("category_hint") or ""),
        document_text=str(value.get("document_text") or ""),
        text_quality_score=float(value.get("text_quality_score") or 0),
        needs_ocr=bool(value.get("needs_ocr") or False),
        needs_vision=bool(value.get("needs_vision") or False),
        pages=pages,
        tables=tables,
        images=images,
        metadata=dict(value.get("metadata") or {}),
        extraction_quality=dict(value.get("extraction_quality") or {}),
        warnings=list(value.get("warnings") or []),
    )


def _ingest_pdf(
    path: Path,
    *,
    vendor_hint: str,
    max_pages: int | None,
    allow_ocr: bool,
) -> DocumentCandidate:
    warnings: list[str] = []
    pages: list[PageCandidate] = []
    tables: list[TableCandidate] = []
    page_limit = max(1, int(max_pages or getattr(settings, "AI_MAX_PAGES", 5) or 5))

    try:
        import pdfplumber  # type: ignore

        with pdfplumber.open(path) as pdf:
            for idx, page in enumerate(pdf.pages[:page_limit], start=1):
                text = page.extract_text() or ""
                blocks: list[TextBlockCandidate] = []
                try:
                    for word in page.extract_words()[:250]:
                        bbox = _normalize_bbox(
                            word.get("x0"),
                            word.get("top"),
                            word.get("x1"),
                            word.get("bottom"),
                            page.width,
                            page.height,
                        )
                        blocks.append(TextBlockCandidate(text=word.get("text", ""), bbox=bbox, source="pdf_text"))
                except Exception:
                    blocks = []
                pages.append(
                    PageCandidate(
                        page_number=idx,
                        text=text,
                        width=float(page.width or 0) or None,
                        height=float(page.height or 0) or None,
                        blocks=blocks,
                    )
                )
                try:
                    for table_index, table in enumerate(page.extract_tables() or [], start=1):
                        normalized = _table_from_rows(
                            table,
                            source="pdf",
                            source_file=path.name,
                            page_number=idx,
                            table_index=table_index,
                            confidence=0.62,
                        )
                        if normalized.rows or normalized.columns:
                            tables.append(normalized)
                except Exception:
                    warnings.append(f"pdf_table_extract_failed_page_{idx}")
    except Exception as exc:
        warnings.append(f"pdf_text_extract_failed:{type(exc).__name__}")

    document_text = _join_page_text(pages)
    source_type = "pdf_digital" if len(document_text.strip()) >= PDF_TEXT_MIN_CHARS else "pdf_scanned"
    if source_type == "pdf_scanned":
        if not allow_ocr:
            warnings.append("pdf_ocr_skipped")
            ocr_pages: list[str] = []
        else:
            ocr_pages = _ocr_pdf_pages(path, page_limit=page_limit, warnings=warnings)
        if any(text.strip() for text in ocr_pages):
            for page_index, ocr_text in enumerate(ocr_pages, start=1):
                if page_index > len(pages):
                    pages.append(
                        PageCandidate(
                            page_number=page_index,
                            text=ocr_text,
                            ocr_confidence=None,
                        )
                    )
                    continue
                if ocr_text.strip():
                    pages[page_index - 1].text = (
                        pages[page_index - 1].text + "\n" + ocr_text
                    ).strip()
            document_text = _join_page_text(pages)
        else:
            warnings.append("pdf_ocr_unavailable_or_empty")

    metadata = _base_metadata(path)
    metadata["page_count_sampled"] = len(pages)
    return DocumentCandidate(
        source_file=path.name,
        source_type=source_type,
        vendor_hint=vendor_hint,
        document_text=_truncate(document_text),
        pages=pages,
        tables=tables,
        metadata=metadata,
        warnings=_dedupe(warnings),
    )


def _ingest_image(path: Path, *, vendor_hint: str, allow_ocr: bool) -> DocumentCandidate:
    warnings: list[str] = []
    width = height = None
    try:
        from PIL import Image  # type: ignore

        with Image.open(path) as img:
            width, height = img.size
    except Exception as exc:
        warnings.append(f"image_metadata_failed:{type(exc).__name__}")

    if allow_ocr:
        text = _ocr_image(path, warnings=warnings)
    else:
        text = ""
        warnings.append("image_ocr_skipped")
    source_type = "screenshot" if _looks_like_screenshot_name(path.name) else "image"
    page = PageCandidate(
        page_number=1,
        text=text,
        image_ref=path.name,
        image_path=str(path),
        width=float(width) if width else None,
        height=float(height) if height else None,
    )
    return DocumentCandidate(
        source_file=path.name,
        source_type=source_type,
        vendor_hint=vendor_hint,
        document_text=_truncate(text),
        pages=[page],
        images=[
            ImageCandidate(
                source_file=path.name,
                page_number=1,
                image_path=str(path),
                width=float(width) if width else None,
                height=float(height) if height else None,
                purpose="screenshot" if source_type == "screenshot" else "original_image",
            )
        ],
        metadata=_base_metadata(path),
        warnings=_dedupe(warnings),
    )


def _ingest_excel(path: Path, *, vendor_hint: str) -> DocumentCandidate:
    warnings: list[str] = []
    if _is_internal_template(path):
        return DocumentCandidate(
            source_file=path.name,
            source_type="internal_template",
            vendor_hint=vendor_hint,
            metadata=_base_metadata(path),
            warnings=["internal_resman_template_not_ingested"],
            extraction_quality={"unsupported_reason": "Internal ResMan template file is not treated as an invoice source."},
        )
    if path.suffix.lower() == ".xls":
        return DocumentCandidate(
            source_file=path.name,
            source_type="excel",
            vendor_hint=vendor_hint,
            metadata=_base_metadata(path),
            warnings=["legacy_xls_not_supported"],
        )
    try:
        import openpyxl  # type: ignore
    except Exception:
        return DocumentCandidate(
            source_file=path.name,
            source_type="excel",
            vendor_hint=vendor_hint,
            metadata=_base_metadata(path),
            warnings=["openpyxl_not_available"],
        )

    tables: list[TableCandidate] = []
    text_parts: list[str] = []
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        cell_count = 0
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows: list[list[Any]] = []
            for row in ws.iter_rows(values_only=True):
                cell_count += len(row)
                if cell_count > MAX_EXCEL_CELLS:
                    warnings.append("excel_workbook_truncated")
                    break
                values = [_cell_text(value) for value in row[:MAX_TABLE_COLS]]
                if any(v != "" for v in values):
                    rows.append(values)
                if len(rows) >= MAX_TABLE_ROWS:
                    warnings.append(f"excel_sheet_truncated:{sheet_name}")
                    break
            table = _table_from_rows(
                rows,
                source="excel",
                source_file=path.name,
                sheet_name=sheet_name,
                table_index=1,
                confidence=0.82,
            )
            if table.rows or table.columns:
                tables.append(table)
                text_parts.append(_table_to_text(sheet_name, table))
    except Exception as exc:
        warnings.append(f"excel_extract_failed:{type(exc).__name__}")

    return DocumentCandidate(
        source_file=path.name,
        source_type="excel",
        vendor_hint=vendor_hint,
        document_text=_truncate("\n\n".join(text_parts)),
        tables=tables,
        metadata=_base_metadata(path),
        warnings=_dedupe(warnings),
    )


def _ingest_word(path: Path, *, vendor_hint: str) -> DocumentCandidate:
    warnings: list[str] = []
    if path.suffix.lower() == ".doc":
        return DocumentCandidate(
            source_file=path.name,
            source_type="word",
            vendor_hint=vendor_hint,
            metadata=_base_metadata(path),
            warnings=["legacy_doc_not_supported"],
        )

    paragraphs: list[str] = []
    tables: list[TableCandidate] = []
    try:
        with zipfile.ZipFile(path) as zf:
            xml = zf.read("word/document.xml")
        root = ET.fromstring(xml)
        ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
        for para in root.findall(".//w:p", ns):
            text = "".join(t.text or "" for t in para.findall(".//w:t", ns)).strip()
            if text:
                paragraphs.append(text)
        for table_index, table in enumerate(root.findall(".//w:tbl", ns), start=1):
            rows: list[list[str]] = []
            for tr in table.findall(".//w:tr", ns):
                row: list[str] = []
                for tc in tr.findall("./w:tc", ns):
                    row.append(" ".join(t.text or "" for t in tc.findall(".//w:t", ns)).strip())
                if any(row):
                    rows.append(row)
            candidate = _table_from_rows(
                rows,
                source="word",
                source_file=path.name,
                sheet_name=f"Table {table_index}",
                table_index=table_index,
                confidence=0.75,
            )
            if candidate.rows or candidate.columns:
                tables.append(candidate)
    except KeyError:
        warnings.append("docx_document_xml_missing")
    except Exception as exc:
        warnings.append(f"docx_extract_failed:{type(exc).__name__}")

    table_text = [_table_to_text(t.sheet_name or "Table", t) for t in tables]
    return DocumentCandidate(
        source_file=path.name,
        source_type="word",
        vendor_hint=vendor_hint,
        document_text=_truncate("\n".join(paragraphs + table_text)),
        tables=tables,
        metadata=_base_metadata(path),
        warnings=_dedupe(warnings),
    )


def _ingest_csv(path: Path, *, vendor_hint: str) -> DocumentCandidate:
    warnings: list[str] = []
    rows: list[list[str]] = []
    try:
        with _open_text_lenient(path) as fh:
            sample = fh.read(4096)
            fh.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample)
            except Exception:
                dialect = csv.excel
            reader = csv.reader(fh, dialect)
            for idx, row in enumerate(reader):
                if idx >= MAX_TABLE_ROWS:
                    warnings.append("csv_truncated")
                    break
                rows.append([str(cell) for cell in row[:MAX_TABLE_COLS]])
    except Exception as exc:
        warnings.append(f"csv_extract_failed:{type(exc).__name__}")
    table = _table_from_rows(
        rows,
        source="csv",
        source_file=path.name,
        sheet_name=path.name,
        table_index=1,
        confidence=0.85,
    )
    return DocumentCandidate(
        source_file=path.name,
        source_type="csv",
        vendor_hint=vendor_hint,
        document_text=_truncate(_table_to_text(path.name, table)),
        tables=[table] if table.rows or table.columns else [],
        metadata=_base_metadata(path),
        warnings=_dedupe(warnings),
    )


def _ingest_unknown(path: Path, *, vendor_hint: str) -> DocumentCandidate:
    warnings = ["unsupported_file_type"]
    text = ""
    if path.suffix.lower() == ".txt":
        try:
            text = path.read_text(encoding="utf-8", errors="replace")
            warnings = ["plain_text_ingested_as_unknown"]
        except Exception as exc:
            warnings.append(f"text_extract_failed:{type(exc).__name__}")
    return DocumentCandidate(
        source_file=path.name,
        source_type="unknown",
        vendor_hint=vendor_hint,
        document_text=_truncate(text),
        metadata=_base_metadata(path),
        warnings=_dedupe(warnings),
    )


def _quick_pdf_source_type(path: Path) -> str:
    try:
        import pdfplumber  # type: ignore

        with pdfplumber.open(path) as pdf:
            page = pdf.pages[0] if pdf.pages else None
            text = page.extract_text() if page else ""
            return "pdf_digital" if len((text or "").strip()) >= 40 else "pdf_scanned"
    except Exception:
        return "pdf_scanned"


def _source_type_from_extension(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        return "pdf_scanned"
    if suffix in IMAGE_EXTENSIONS:
        return "screenshot" if _looks_like_screenshot_name(path.name) else "image"
    if suffix in EXCEL_EXTENSIONS:
        return "excel"
    if suffix in WORD_EXTENSIONS:
        return "word"
    if suffix == ".csv":
        return "csv"
    return "unknown"


def _source_type_label(source_type: str) -> str:
    return {
        "pdf_digital": "PDF digital",
        "pdf_scanned": "PDF scanned",
        "image": "Image",
        "screenshot": "Screenshot",
        "excel": "Excel",
        "word": "Word",
        "csv": "CSV",
        "internal_template": "Internal template ignored",
        "unknown": "Unsupported",
    }.get(source_type, "Unsupported")


def _ocr_pdf_pages(path: Path, *, page_limit: int, warnings: list[str]) -> list[str]:
    """Return OCR text aligned to source PDF pages.

    Keeping page boundaries is essential for PDFs that contain several
    independent invoices. The former implementation returned one concatenated
    string and assigned it to page 1, which made downstream invoice grouping
    impossible and left the remaining page candidates blank.
    """
    cache_dpi = 180
    try:
        from utils import ocr_cache  # type: ignore
    except Exception:
        ocr_cache = None  # type: ignore
    if ocr_cache is not None:
        try:
            cached = ocr_cache.lookup(path, cache_dpi)
        except Exception:
            cached = None
        if cached and int(cached.get("page_limit") or 0) >= page_limit:
            page_texts = [
                str(page.get("text") or "").strip()
                for page in (cached.get("pages") or [])[:page_limit]
                if isinstance(page, dict)
            ]
            if any(page_texts):
                return page_texts
    try:
        import pypdfium2 as pdfium  # type: ignore
        from PIL import ImageEnhance, ImageOps  # type: ignore
        import pytesseract  # type: ignore
    except Exception:
        warnings.append("pdf_ocr_dependencies_unavailable")
        return []
    try:
        doc = pdfium.PdfDocument(str(path))
    except Exception as exc:
        warnings.append(f"pdf_render_failed:{type(exc).__name__}")
        return []
    texts: list[str] = []
    cached_pages: list[dict[str, Any]] = []
    try:
        for page_index in range(min(len(doc), page_limit)):
            page = doc[page_index]
            width = float(page.get_width() or 612)
            scale = min(3.0, max(2.0, 1800 / width))
            img = page.render(scale=scale).to_pil()
            img = ImageOps.grayscale(img)
            img = ImageEnhance.Contrast(img).enhance(1.7)
            try:
                text = pytesseract.image_to_string(img, config="--psm 6").strip()
            except Exception:
                text = ""
            texts.append(text)
            cached_pages.append({
                "page_number": page_index + 1,
                "text": text,
                "width": float(img.width),
                "height": float(img.height),
                "words": [],
            })
    finally:
        try:
            doc.close()
        except Exception:
            pass
    if ocr_cache is not None and any(texts):
        try:
            ocr_cache.store(path, cache_dpi, {
                "pages": cached_pages,
                "page_limit": page_limit,
                "extraction_method": "pdf_ocr",
                "confidence": 0.0,
                "warnings": [],
            })
        except Exception:
            pass
    return texts


def _ocr_pdf(path: Path, *, page_limit: int, warnings: list[str]) -> str:
    """Backward-compatible concatenated OCR text helper."""
    return "\n\n".join(
        text for text in _ocr_pdf_pages(path, page_limit=page_limit, warnings=warnings)
        if text
    )


def _ocr_image(path: Path, *, warnings: list[str]) -> str:
    # Phase PERF-1 hotfix — cache hit. _ocr_image runs five Tesseract
    # variants on the same image to score the best result; that's
    # ~10–50 seconds for a single screenshot. Caching by file hash
    # eliminates the cost on every re-detect / re-process for the
    # exact same bytes.
    try:
        from utils import ocr_cache  # type: ignore
    except Exception:
        ocr_cache = None  # type: ignore
    if ocr_cache is not None:
        try:
            cached = ocr_cache.lookup(path, 0)  # dpi=0 for raw-image cache
        except Exception:
            cached = None
        if cached:
            pages = cached.get("pages") or []
            if pages:
                warnings.append("image_ocr_cache_hit")
                return pages[0].get("text") or ""
    try:
        from PIL import Image, ImageEnhance, ImageFilter, ImageOps  # type: ignore
        import pytesseract  # type: ignore
    except Exception:
        warnings.append("image_ocr_dependencies_unavailable")
        return ""
    try:
        with Image.open(path) as img:
            base = ImageOps.grayscale(img)
            variants = []
            variants.append(("gray_contrast_psm6", ImageEnhance.Contrast(base).enhance(1.8), "--psm 6"))
            variants.append(("gray_contrast_psm11", ImageEnhance.Contrast(base).enhance(1.9), "--psm 11"))
            upscaled = base
            if min(base.size or (0, 0)) < 1400:
                upscaled = base.resize((base.width * 2, base.height * 2))
            sharp = ImageEnhance.Sharpness(ImageEnhance.Contrast(upscaled).enhance(2.0)).enhance(1.6)
            variants.append(("upscale_sharp_psm6", sharp, "--psm 6"))
            variants.append(("upscale_sharp_psm11", sharp, "--psm 11"))
            threshold = sharp.point(lambda p: 255 if p > 170 else 0)
            variants.append(("threshold_psm6", threshold.filter(ImageFilter.MedianFilter(size=3)), "--psm 6"))

            best_text = ""
            best_score = -1.0
            best_variant = ""
            for variant_name, image, config in variants:
                try:
                    text = pytesseract.image_to_string(image, config=config).strip()
                except Exception:
                    continue
                score = _text_quality_score(text)
                if score > best_score or (score == best_score and len(text) > len(best_text)):
                    best_score = score
                    best_text = text
                    best_variant = variant_name
            if best_variant:
                warnings.append(f"image_ocr_variant:{best_variant}")
            # Phase PERF-1 hotfix — persist for the next re-detect.
            if ocr_cache is not None and best_text:
                try:
                    ocr_cache.store(path, 0, {
                        "pages": [{
                            "page_number": 1,
                            "text": best_text,
                            "width": 0,
                            "height": 0,
                            "words": [],
                        }],
                        "extraction_method": "image_ocr",
                        "confidence": 0.0,
                        "warnings": [w for w in warnings if w.startswith("image_ocr_")],
                    })
                except Exception:
                    pass
            return best_text
    except Exception as exc:
        warnings.append(f"image_ocr_failed:{type(exc).__name__}")
        return ""


def _table_from_rows(
    rows: list[list[Any]],
    *,
    source: str,
    source_file: str = "",
    sheet_name: str | None = None,
    page_number: int | None = None,
    table_index: int = 1,
    confidence: float | None = None,
) -> TableCandidate:
    cleaned = [[_cell_text(cell) for cell in row[:MAX_TABLE_COLS]] for row in rows[:MAX_TABLE_ROWS]]
    cleaned = [row for row in cleaned if any(cell != "" for cell in row)]
    columns = cleaned[0] if cleaned else []
    data = cleaned[1:] if len(cleaned) > 1 else []
    return TableCandidate(
        source=source,
        source_file=source_file,
        source_type=source,
        sheet_name=sheet_name,
        page_number=page_number,
        table_index=table_index,
        headers=columns,
        rows=data,
        columns=columns,
        confidence=confidence,
        raw_cells=cleaned,
    )


def _table_to_text(label: str, table: TableCandidate) -> str:
    lines = [f"[{label}]"]
    if table.columns:
        lines.append(" | ".join(str(v) for v in table.columns))
    for row in table.rows[:80]:
        lines.append(" | ".join(str(v) for v in row))
    return "\n".join(lines)


def _quality(candidate: DocumentCandidate) -> dict[str, Any]:
    text = candidate.document_text or ""
    table_count = len(candidate.tables)
    image_count = len(candidate.images)
    quality = _text_quality_score(text)
    field_evidence = _invoice_field_evidence(text)
    accounting_completeness = float(field_evidence["score"])
    quality_label = _quality_label(quality, has_tables=table_count > 0)
    needs_vision = (
        candidate.source_type in {"image", "screenshot", "pdf_scanned"}
        and (quality < 0.45 or accounting_completeness < 0.72)
    )
    return {
        "label": quality_label,
        "text_chars": len(text),
        "page_count": len(candidate.pages),
        "table_count": table_count,
        "image_count": image_count,
        "text_quality_score": quality,
        "accounting_completeness": accounting_completeness,
        "critical_fields_found": field_evidence["found"],
        "critical_fields_missing": field_evidence["missing"],
        "needs_ocr": candidate.source_type in {"pdf_scanned", "image", "screenshot"} and not text.strip(),
        "vision_recommended": needs_vision,
        "warnings_count": len(candidate.warnings),
    }


def _finalize_candidate(
    candidate: DocumentCandidate,
    path: Path,
    *,
    allow_vision_hint: bool,
) -> DocumentCandidate:
    base = _base_metadata(path)
    quality = _quality(candidate)
    initial_quality = dict(candidate.extraction_quality or {})
    initial_quality.update(quality)
    if not allow_vision_hint:
        initial_quality["vision_recommended"] = False

    candidate.source_file = candidate.source_file or path.name
    candidate.source_path = candidate.source_path or path.name
    candidate.mime_type = candidate.mime_type or base["mime_type"]
    candidate.file_size_bytes = candidate.file_size_bytes or int(base["size_bytes"] or 0)
    candidate.page_count = candidate.page_count or _candidate_page_count(path, candidate)
    candidate.sheet_count = candidate.sheet_count or _candidate_sheet_count(candidate)
    candidate.text_quality_score = float(quality.get("text_quality_score") or 0)
    candidate.needs_ocr = bool(quality.get("needs_ocr") or False)
    candidate.needs_vision = bool(quality.get("vision_recommended") or False) and allow_vision_hint
    candidate.extraction_quality = initial_quality
    candidate.metadata = {**base, **dict(candidate.metadata or {})}
    candidate.metadata.setdefault("page_count", candidate.page_count)
    candidate.metadata.setdefault("sheet_count", candidate.sheet_count)
    candidate.metadata.setdefault("quality_label", quality.get("label"))

    if not candidate.vendor_hint:
        candidate.vendor_hint = detect_vendor_hint(candidate)
    if not candidate.category_hint:
        candidate.category_hint = detect_category_hint(candidate)

    for page in candidate.pages:
        page.text_quality_score = _text_quality_score(page.text or "")
        if page.image_path and not page.image_ref:
            page.image_ref = Path(page.image_path).name

    if candidate.text_quality_score == 0:
        candidate.warnings.append("no_text_extracted")
    elif candidate.text_quality_score < 0.45:
        candidate.warnings.append("weak_text_quality")
    if candidate.needs_vision:
        candidate.warnings.append("vision_recommended")
    candidate.warnings = _dedupe(candidate.warnings)
    candidate.extraction_quality["warnings_count"] = len(candidate.warnings)
    return candidate


def _quality_label(score: float, *, has_tables: bool) -> str:
    if score >= 0.70 or (has_tables and score >= 0.55):
        return "high"
    if score >= 0.45 or has_tables:
        return "medium"
    if score > 0:
        return "low"
    return "none"


def _text_quality_score(text: str) -> float:
    normalized = re.sub(r"\s+", " ", text or "").strip()
    if not normalized:
        return 0.0
    lowered = normalized.lower()
    tokens = re.findall(r"[a-z0-9$.,/#-]+", lowered)
    alnum = sum(ch.isalnum() for ch in normalized)
    alnum_ratio = alnum / max(1, len(normalized))

    # Character density alone is a poor OCR signal: a page full of plausible
    # looking nonsense still has a high alphanumeric ratio. Require invoice
    # semantics and structured values before calling scanned text usable.
    label_terms = (
        "invoice", "account", "total", "amount", "date", "due",
        "description", "quantity", "price", "subtotal", "tax",
    )
    label_hits = sum(1 for term in label_terms if re.search(rf"\b{term}\b", lowered))
    has_money = bool(re.search(r"(?:\$\s*)?\d{1,3}(?:,\d{3})*(?:\.\d{2})\b", normalized))
    has_date = bool(
        re.search(r"\b\d{1,2}[/-]\d{1,2}[/-](?:\d{2}|\d{4})\b", normalized)
        or re.search(
            r"\b(?:jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*\s+\d{1,2},?\s+\d{4}\b",
            lowered,
        )
    )
    has_invoice_id = bool(
        re.search(r"\binvoice\s*(?:#|no\.?|number)?\s*[:#-]?\s*[a-z0-9-]{2,}\b", lowered)
    )
    short_noise_ratio = (
        sum(1 for token in tokens if len(re.sub(r"[^a-z0-9]", "", token)) <= 2)
        / max(1, len(tokens))
    )

    score = min(0.38, alnum_ratio * 0.45)
    score += min(0.30, label_hits * 0.05)
    score += 0.12 if has_money else 0.0
    score += 0.10 if has_date else 0.0
    score += 0.10 if has_invoice_id else 0.0
    if len(tokens) < 12:
        score -= 0.10
    if short_noise_ratio > 0.30:
        score -= min(0.18, (short_noise_ratio - 0.30) * 0.60)
    evidence = _invoice_field_evidence(normalized)
    missing = set(evidence["missing"])
    if {"invoice_number", "total_amount"}.issubset(missing):
        score = min(score, 0.54)
    elif {"invoice_number", "total_amount"} & missing:
        score = min(score, 0.68)
    return round(max(0.0, min(1.0, score)), 3)


def _invoice_field_evidence(text: str) -> dict[str, Any]:
    """Measure whether OCR captured values, not merely invoice labels."""
    lines = [re.sub(r"\s+", " ", line).strip() for line in (text or "").splitlines()]
    lines = [line for line in lines if line]
    lowered = [line.lower() for line in lines]

    invoice_number = False
    for index, line in enumerate(lowered):
        match = re.search(r"\binvoice\s*(?:number|no\.?|#)\s*[:#-]?\s*(.*)$", line)
        if not match:
            continue
        candidates = [match.group(1)]
        if index + 1 < len(lowered) and not re.search(
            r"\b(?:sold|bill|ship)\s+to\b", lowered[index + 1],
        ):
            candidates.append(lowered[index + 1])
        for candidate in candidates:
            candidate = re.split(
                r"\b(?:account|sales|location|invoice\s+date|date|terms)\b",
                candidate,
                maxsplit=1,
            )[0]
            if re.search(r"\b(?=[a-z0-9-]*\d)[a-z0-9-]{3,}\b", candidate):
                invoice_number = True
                break
        if invoice_number:
            break

    date_found = bool(re.search(
        r"\b\d{1,2}[/-]\d{1,2}[/-](?:\d{2}|\d{4})\b",
        "\n".join(lines),
    ))
    money_pattern = r"(?:\$\s*)?\d{1,3}(?:,\d{3})*(?:\.\d{2})\b"
    money_values = re.findall(money_pattern, "\n".join(lines))
    total_found = any(
        re.search(r"\b(?:grand\s+total|invoice\s+total|total\s+due|amount\s+due|total)\b", line)
        and re.search(money_pattern, line)
        for line in lowered
    )
    line_amount = len(money_values) >= 1
    vendor_identity = any(
        len(re.findall(r"[a-z]{3,}", line)) >= 2
        and not re.search(r"\b(?:invoice|account|description|quantity|subtotal|total)\b", line)
        for line in lowered[:8]
    )
    property_evidence = bool(
        re.search(r"\b(?:sold|bill|ship|service|install)\s+to\b", "\n".join(lowered))
        or re.search(r"[a-z0-9._%+-]+@[a-z0-9.-]+\.[a-z]{2,}", "\n".join(lowered))
    )
    found_map = {
        "vendor_name": vendor_identity,
        "invoice_number": invoice_number,
        "invoice_date": date_found,
        "total_amount": total_found,
        "line_amounts": line_amount,
        "property_context": property_evidence,
    }
    weights = {
        "vendor_name": 0.15,
        "invoice_number": 0.20,
        "invoice_date": 0.15,
        "total_amount": 0.25,
        "line_amounts": 0.15,
        "property_context": 0.10,
    }
    return {
        "score": round(sum(weights[key] for key, found in found_map.items() if found), 3),
        "found": [key for key, found in found_map.items() if found],
        "missing": [key for key, found in found_map.items() if not found],
    }


def detect_vendor_hint(candidate: DocumentCandidate) -> str:
    text = (candidate.document_text or "").lower()
    hints = (
        ("capital waste", "Capital Waste Services"),
        ("lowe", "Lowe's Pro Supply"),
        ("hd supply", "HD Supply"),
        ("tk elevator", "TK Elevator"),
        ("epb", "EPB Fiber Optics"),
        ("spectrum", "Spectrum"),
        ("pennyrile", "Pennyrile Electric"),
        ("weakley", "Weakley County Municipal Electric System"),
        ("richmond utilities", "Richmond Utilities"),
        ("hopkinsville water", "Hopkinsville Water Environment Authority"),
        ("hwea", "Hopkinsville Water Environment Authority"),
        ("servall", "Servall Pest"),
        ("alabama power", "Alabama Power"),
        ("kentucky utilities", "Kentucky Utilities"),
        ("knoxville utilities", "Knoxville Utility Board"),
        ("kub", "Knoxville Utility Board"),
    )
    for needle, label in hints:
        if needle in text:
            return label
    return ""


def detect_category_hint(candidate: DocumentCandidate) -> str:
    text = f"{candidate.vendor_hint} {candidate.document_text}".lower()
    if any(token in text for token in ("waste", "trash", "garbage", "refuse", "dumpster")):
        return "trash_collection_services"
    if any(token in text for token in ("electric", "kwh", "water", "sewer", "wastewater", "gas", "fiber", "internet", "utility")):
        return "utilities"
    if any(token in text for token in ("pest", "termite", "exterminat")):
        return "pest_control"
    if any(token in text for token in ("landscap", "lawn", "mulch", "mowing")):
        return "landscaping"
    if any(token in text for token in ("subscription", "software", "license")):
        return "subscriptions"
    if any(token in text for token in ("advertising", "marketing")):
        return "marketing"
    if text.strip():
        return "other_infrequent"
    return "unknown"


def _candidate_page_count(path: Path, candidate: DocumentCandidate) -> int:
    if candidate.source_type in {"image", "screenshot"}:
        return 1
    if candidate.source_type == "pdf_digital" or candidate.source_type == "pdf_scanned":
        try:
            import pdfplumber  # type: ignore

            with pdfplumber.open(path) as pdf:
                return len(pdf.pages)
        except Exception:
            return len(candidate.pages)
    return len(candidate.pages)


def _candidate_sheet_count(candidate: DocumentCandidate) -> int:
    sheets = {t.sheet_name for t in candidate.tables if t.sheet_name}
    return len(sheets)


def _normalize_bbox(
    x0: Any,
    y0: Any,
    x1: Any,
    y1: Any,
    width: Any,
    height: Any,
) -> dict[str, float] | None:
    try:
        w = float(width or 0)
        h = float(height or 0)
        if w <= 0 or h <= 0:
            return None
        return {
            "x": round(float(x0 or 0) / w, 6),
            "y": round(float(y0 or 0) / h, 6),
            "w": round((float(x1 or 0) - float(x0 or 0)) / w, 6),
            "h": round((float(y1 or 0) - float(y0 or 0)) / h, 6),
        }
    except Exception:
        return None


def _join_page_text(pages: list[PageCandidate]) -> str:
    return "\n\n".join(page.text for page in pages if page.text)


def _truncate(text: str) -> str:
    return (text or "")[: max(1000, int(getattr(settings, "AI_MAX_TEXT_CHARS", 45000) or 45000))]


def _base_metadata(path: Path) -> dict[str, Any]:
    try:
        stat = path.stat()
        size = stat.st_size
        mtime = int(stat.st_mtime)
    except OSError:
        size = 0
        mtime = 0
    return {
        "extension": path.suffix.lower(),
        "size_bytes": size,
        "mtime": mtime,
        "mime_type": mimetypes.guess_type(path.name)[0] or "application/octet-stream",
    }


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except Exception:
            pass
    return str(value).strip()


def _open_text_lenient(path: Path) -> StringIO:
    raw = path.read_bytes()
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return StringIO(raw.decode(enc))
        except UnicodeDecodeError:
            continue
    return StringIO(raw.decode("latin-1", errors="replace"))


def _looks_like_screenshot_name(name: str) -> bool:
    lowered = name.lower()
    return any(token in lowered for token in ("screenshot", "screen shot", "cam scanner", "camscanner", "scan"))


def _is_internal_template(path: Path) -> bool:
    try:
        resolved = path.resolve()
        if resolved == settings.RESMAN_TEMPLATE.resolve():
            return True
        lower_parts = [part.lower() for part in resolved.parts]
        return path.name.lower() == "template.xlsx" and "output" in lower_parts
    except Exception:
        return False


def _dedupe(values: list[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for value in values:
        if value and value not in seen:
            seen.add(value)
            out.append(value)
    return out


__all__ = [
    "DocumentCandidate",
    "ImageCandidate",
    "PageCandidate",
    "TableCandidate",
    "TextBlockCandidate",
    "detect_category_hint",
    "detect_file_support",
    "detect_source_type",
    "detect_vendor_hint",
    "document_candidate_from_dict",
    "ingest_document",
]
