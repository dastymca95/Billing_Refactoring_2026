"""ResMan template column metadata.

Combines two sources of truth:
  * `Output/Template.xlsx` — canonical column ORDER and the exact header text.
  * `config/resman_template_rules.yaml` — which columns are required vs
    recommended vs optional, plus a few UI behavior flags.

Returns a small dict the frontend can use to render the preview table:

    {
      "columns": ["Invoice Number", ..., "Document Url"],
      "required_columns": [...],
      "recommended_columns": [...],
      "optional_columns": [...],
      "optional_columns_collapsible": True,
      "optional_columns_hidden_by_default": True,
    }

The reader caches the result for the lifetime of the process; the YAML and
Template.xlsx don't change at runtime, so a single read is fine.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

import yaml
import openpyxl

from ..settings import PROJECT_ROOT, RESMAN_TEMPLATE


_TEMPLATE_RULES_YAML = PROJECT_ROOT / "config" / "resman_template_rules.yaml"
_CANONICAL_RULES_YAML = PROJECT_ROOT / "config" / "canonical_rules.yaml"
_FORMAT_RULES_YAML = PROJECT_ROOT / "config" / "invoice_format_rules.yaml"

_LOG = logging.getLogger("template_rules")
_CACHED: dict[str, Any] | None = None


def _read_template_columns(template_path: Path, header_row: int) -> list[str]:
    """Read the header row of `Output/Template.xlsx` and return the column
    names in their canonical order, dropping trailing empty cells."""
    wb = openpyxl.load_workbook(template_path, read_only=True, data_only=True)
    sheet = wb.active
    if "Sheet 1" in wb.sheetnames:
        sheet = wb["Sheet 1"]
    cols: list[str] = []
    for c in range(1, sheet.max_column + 1):
        v = sheet.cell(row=header_row, column=c).value
        if v is None:
            continue
        s = str(v).strip()
        if s:
            cols.append(s)
    wb.close()
    return cols


def get_template_rules() -> dict[str, Any]:
    """Return the cached template-rules metadata. Falls back to a sensible
    default if the YAML or template can't be read."""
    global _CACHED
    if _CACHED is not None:
        return _CACHED

    cfg: dict[str, Any] = {}
    if _TEMPLATE_RULES_YAML.is_file():
        try:
            cfg = yaml.safe_load(_TEMPLATE_RULES_YAML.read_text(encoding="utf-8")) or {}
        except Exception as e:  # pragma: no cover
            _LOG.warning("Could not load %s: %s", _TEMPLATE_RULES_YAML, e)
            cfg = {}

    template_block = cfg.get("template") or {}
    template_path = Path(template_block.get("path") or "")
    if not template_path.is_absolute():
        template_path = (PROJECT_ROOT / template_path) if template_path else RESMAN_TEMPLATE
    if not template_path.is_file():
        template_path = RESMAN_TEMPLATE

    header_row = int(template_block.get("header_row") or 1)

    try:
        columns = _read_template_columns(template_path, header_row)
    except Exception as e:
        _LOG.warning("Could not read columns from %s: %s", template_path, e)
        # Conservative fallback so the app still boots.
        columns = [
            "Invoice Number", "Bill or Credit", "Invoice Date", "Accounting Date",
            "Vendor", "Invoice Description", "Line Item Number",
            "Property Abbreviation", "Location", "GL Account",
            "Line Item Description", "Amount", "Expense Type",
            "Is Replacement Reserve", "Payment Date", "Reference Number",
            "Payment Method", "Department", "Due Date", "Quantity",
            "Unit Price", "Tax", "Received Date", "Document Url",
        ]

    canonical_requirements = _read_canonical_required_columns()
    format_requirements = _read_format_required_columns()
    configured_required = canonical_requirements or format_requirements or cfg.get("required_columns") or []
    required = [c for c in configured_required if c in columns]
    recommended = [
        c for c in (cfg.get("recommended_columns") or [])
        if c in columns and c not in required
    ]
    classified = set(required) | set(recommended)
    optional = [c for c in columns if c not in classified]

    _CACHED = {
        "columns": columns,
        "required_columns": required,
        "recommended_columns": recommended,
        "optional_columns": optional,
        "optional_columns_collapsible": bool(
            cfg.get("optional_columns_collapsible", True)
        ),
        "optional_columns_hidden_by_default": bool(
            cfg.get("optional_columns_hidden_by_default", True)
        ),
        "template_path": str(template_path),
    }
    return _CACHED


def reset_cache() -> None:
    """Force the next call to re-read the YAML + template (used by tests)."""
    global _CACHED
    _CACHED = None


def _read_format_required_columns() -> list[str]:
    """Return operator-configured export-required columns from Formats.

    `resman_template_rules.yaml` remains the fallback/canonical default, but
    the app's Formats screen is the operator-facing source for which template
    fields are truly mandatory before export.
    """
    if not _FORMAT_RULES_YAML.is_file():
        return []
    try:
        data = yaml.safe_load(_FORMAT_RULES_YAML.read_text(encoding="utf-8")) or {}
    except Exception:
        return []
    requirements = data.get("template_requirements")
    if not isinstance(requirements, dict):
        return []
    required = requirements.get("required_columns")
    if not isinstance(required, list):
        return []
    out: list[str] = []
    seen: set[str] = set()
    for item in required:
        text = str(item or "").strip()
        if not text or text in seen:
            continue
        seen.add(text)
        out.append(text)
    return out


def _read_canonical_required_columns() -> list[str]:
    """Return export-required columns from Canonical Rules when available."""
    if not _CANONICAL_RULES_YAML.is_file():
        return []
    try:
        data = yaml.safe_load(_CANONICAL_RULES_YAML.read_text(encoding="utf-8")) or {}
    except Exception:
        return []
    requirements = data.get("template_requirements")
    if not isinstance(requirements, dict):
        return []
    required = requirements.get("required_columns")
    if not isinstance(required, list):
        return []
    out: list[str] = []
    seen: set[str] = set()
    for item in required:
        text = str(item or "").strip()
        if not text or text in seen:
            continue
        seen.add(text)
        out.append(text)
    return out
