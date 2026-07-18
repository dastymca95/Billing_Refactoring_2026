"""Batch processing facade — routes each batch to the right vendor processor.

Phase 1 supports Richmond Utilities only. The facade reads the detection
results, finds files that belong to a supported vendor, calls the
processor, and returns a unified result the API layer can return as JSON.
"""

from __future__ import annotations

import importlib
import json
import logging
import os
import shutil
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Optional
from urllib.parse import quote, urlparse

import openpyxl

from ..settings import (
    PROJECT_ROOT, RESMAN_TEMPLATE, VENDORS_DIR, batch_dir,
)
from . import batch_store
from .template_rules import get_template_rules
from .vendor_detection import (
    detect_vendor_for_file,
    detect_vendor_from_text,
    fast_detection_context,
)


_LOG = logging.getLogger(__name__)


def _supported_file_count(files: list[Path], unsupported: list[dict[str, Any]]) -> int:
    """Count fully supported source files without confusing attempts with success."""
    failed_source_files = {
        str(item.get("filename") or "").strip()
        for item in unsupported
        if isinstance(item, dict) and str(item.get("filename") or "").strip()
    }
    return max(0, len(files) - len(failed_source_files))


# Phase 1H — declared timeline stages. The batch processor declares
# these up-front so the frontend can render the full list as
# `pending` placeholders before any work begins. Stage keys are stable
# and used by both the processor and the frontend renderer.
_DEFAULT_STAGES: list[tuple[str, str]] = [
    ("upload", "Uploading files"),
    ("vendor_detect", "Detecting vendor"),
    ("read_pdf", "Reading PDF text"),
    ("ocr", "Running OCR"),
    ("yaml_rules", "Applying vendor YAML rules"),
    ("address_match", "Matching service address"),
    ("unit_match", "Matching Unit Info Clean"),
    ("gl_evidence", "Using General Ledger evidence"),
    ("ai_fallback", "AI invoice assist (if enabled)"),
    ("reconcile", "Reconciling bill totals"),
    ("split_pdf", "Splitting support PDFs"),
    ("dropbox", "Uploading to Dropbox"),
    ("template", "Building ResMan template"),
    ("ready", "Ready for review"),
]


def _read_batch_metadata(batch_id: str) -> dict:
    """Read the batch_metadata.json sidecar (the same one batches.py
    writes). Returns an empty dict if missing."""
    bdir = batch_store.get_batch_dir(batch_id)
    p = bdir / "batch_metadata.json"
    if not p.is_file():
        return {}
    try:
        return json.loads(p.read_text(encoding="utf-8")) or {}
    except Exception:
        return {}


def _read_region_hints(batch_id: str) -> list[dict]:
    """Read the region_hints.json sidecar (regions.py writes it). Returns
    an empty list if missing or malformed."""
    bdir = batch_store.get_batch_dir(batch_id)
    p = bdir / "region_hints.json"
    if not p.is_file():
        return []
    try:
        data = json.loads(p.read_text(encoding="utf-8")) or {}
        return list(data.get("regions") or [])
    except Exception:
        return []


# ---------------------------------------------------------------------------
# Vendor → processor module + entrypoint name. As new vendors come online,
# add a loader function below + register it in _PROCESSOR_LOADERS. Each
# loader returns the processor module; the registry tells `process_batch`
# which top-level function to call inside it.
# ---------------------------------------------------------------------------
def _import_richmond_processor():
    """Import the Richmond Utilities processor module."""
    vendor_folder = (PROJECT_ROOT / "Training Bills_Invoices" / "Water - Sewer"
                     / "Richmond Utilities").resolve()
    if str(vendor_folder) not in sys.path:
        sys.path.insert(0, str(vendor_folder))
    return importlib.import_module("process_richmond_utilities")


def _import_hopkinsville_processor():
    """Import the Hopkinsville Water Environment Authority processor module."""
    vendor_folder = (PROJECT_ROOT / "Training Bills_Invoices" / "Water - Sewer"
                     / "Hopkinsville Water Environment Authority").resolve()
    if str(vendor_folder) not in sys.path:
        sys.path.insert(0, str(vendor_folder))
    return importlib.import_module("process_hopkinsville_water_environment_authority")


def _import_columbia_processor():
    """Import the Columbia Power and Water System processor module."""
    vendor_folder = (PROJECT_ROOT / "Training Bills_Invoices" / "Electricity - Power"
                     / "Columbia Power and Water System").resolve()
    if str(vendor_folder) not in sys.path:
        sys.path.insert(0, str(vendor_folder))
    return importlib.import_module("process_columbia_power_and_water_system")


def _import_atmos_processor():
    """Import the Atmos Energy Auto Pay processor module."""
    vendor_folder = (PROJECT_ROOT / "Training Bills_Invoices" / "Gas"
                     / "Atmos Energy Auto Pay").resolve()
    if str(vendor_folder) not in sys.path:
        sys.path.insert(0, str(vendor_folder))
    return importlib.import_module("process_atmos_energy_auto_pay")


def _import_hardin_processor():
    """Import the Hardin County Water District No. 2 processor module."""
    vendor_folder = (PROJECT_ROOT / "Training Bills_Invoices" / "Water - Sewer"
                     / "Hardin County Water District No. 2").resolve()
    if str(vendor_folder) not in sys.path:
        sys.path.insert(0, str(vendor_folder))
    return importlib.import_module("process_hardin_county_water_district_no_2")


def _import_shelbyville_processor():
    """Import the Shelbyville Power System processor module."""
    vendor_folder = (PROJECT_ROOT / "Training Bills_Invoices" / "Electricity - Power"
                     / "Shelbyville Power System").resolve()
    if str(vendor_folder) not in sys.path:
        sys.path.insert(0, str(vendor_folder))
    return importlib.import_module("process_shelbyville_power_system")


def _import_zillow_processor():
    """Import the webapp-native deterministic Zillow Rentals processor."""
    return importlib.import_module("webapp.backend.services.zillow_processor")


def _import_apartments_com_processor():
    """Import the webapp-native deterministic Apartments.com processor."""
    return importlib.import_module("webapp.backend.services.apartments_com_processor")


def _import_mcminnville_processor():
    """Import the McMinnville Electric System processor module."""
    vendor_folder = (PROJECT_ROOT / "Training Bills_Invoices" / "Electricity - Power"
                     / "McMinnville Electric System").resolve()
    if str(vendor_folder) not in sys.path:
        sys.path.insert(0, str(vendor_folder))
    return importlib.import_module("process_mcminnville_electric_system")


def _import_pennyrile_processor():
    """Import the Pennyrile Electric processor module."""
    vendor_folder = (PROJECT_ROOT / "Training Bills_Invoices" / "Electricity - Power"
                     / "Pennyrile Electric").resolve()
    if str(vendor_folder) not in sys.path:
        sys.path.insert(0, str(vendor_folder))
    return importlib.import_module("process_pennyrile_electric")


def _import_wave2_utility_processor():
    """Import the shared Phase U2 utility processor module."""
    return importlib.import_module("webapp.backend.services.utility_wave2_processors")


def _import_wave3_utility_processor():
    """Import the shared Phase U3 utility processor module."""
    return importlib.import_module("webapp.backend.services.utility_wave3_processors")


def _import_resman_processor():
    """Import the deterministic ResMan, LLC processor module."""
    return importlib.import_module("webapp.backend.services.resman_processor")


def _import_granite_telecommunications_processor():
    """Import the deterministic Granite Telecommunications processor."""
    return importlib.import_module(
        "webapp.backend.services.granite_telecommunications_processor"
    )


def _import_lowes_pro_supply_processor():
    """Import the deterministic Lowe's Pro Supply processor."""
    return importlib.import_module("webapp.backend.services.lowes_pro_supply_processor")


# vendor_key → (loader, entrypoint_name)
_PROCESSOR_LOADERS: dict[str, tuple[Any, str]] = {
    "richmond_utilities": (_import_richmond_processor,
                           "process_richmond_utilities_batch"),
    "hopkinsville_water_environment_authority": (
        _import_hopkinsville_processor,
        "process_hopkinsville_water_environment_authority_batch",
    ),
    "columbia_power_and_water_system": (
        _import_columbia_processor,
        "process_columbia_power_and_water_system_batch",
    ),
    "atmos_energy_auto_pay": (
        _import_atmos_processor,
        "process_atmos_energy_auto_pay_batch",
    ),
    "hardin_county_water_district_no_2": (
        _import_hardin_processor,
        "process_hardin_county_water_district_no_2_batch",
    ),
    "shelbyville_power_system": (
        _import_shelbyville_processor,
        "process_shelbyville_power_system_batch",
    ),
    "zillow_rentals": (
        _import_zillow_processor,
        "process_zillow_rentals_batch",
    ),
    "apartments_com": (
        _import_apartments_com_processor,
        "process_apartments_com_batch",
    ),
    "resman_llc": (
        _import_resman_processor,
        "process_resman_llc_batch",
    ),
    "granite_telecommunications_llc": (
        _import_granite_telecommunications_processor,
        "process_granite_telecommunications_batch",
    ),
    "lowes": (
        _import_lowes_pro_supply_processor,
        "process_lowes_pro_supply_batch",
    ),
    "mcminnville_electric_system": (
        _import_mcminnville_processor,
        "process_mcminnville_electric_system_batch",
    ),
    "pennyrile_electric": (
        _import_pennyrile_processor,
        "process_pennyrile_electric_batch",
    ),
    "alabama_power": (
        _import_wave2_utility_processor,
        "process_alabama_power_batch",
    ),
    "epb_fiber_optics": (
        _import_wave2_utility_processor,
        "process_epb_fiber_optics_batch",
    ),
    "the_city_of_henderson": (
        _import_wave2_utility_processor,
        "process_the_city_of_henderson_batch",
    ),
    "cde_lightband": (
        _import_wave2_utility_processor,
        "process_cde_lightband_batch",
    ),
    "nolin_recc_smarthub": (
        _import_wave2_utility_processor,
        "process_nolin_recc_smarthub_batch",
    ),
    "clarksville_gas_and_water": (
        _import_wave3_utility_processor,
        "process_clarksville_gas_and_water_batch",
    ),
    "knoxville_utilities_board": (
        _import_wave3_utility_processor,
        "process_knoxville_utilities_board_batch",
    ),
    "kentucky_utilities": (
        _import_wave3_utility_processor,
        "process_kentucky_utilities_batch",
    ),
    "tennessee_american_water": (
        _import_wave3_utility_processor,
        "process_tennessee_american_water_batch",
    ),
    "union_city_energy_authority": (
        _import_wave3_utility_processor,
        "process_union_city_energy_authority_batch",
    ),
    "nashville_electric_service": (
        _import_wave3_utility_processor,
        "process_nashville_electric_service_batch",
    ),
    "weakley_county_municipal_electric_system": (
        _import_wave3_utility_processor,
        "process_weakley_county_municipal_electric_system_batch",
    ),
    "birmingham_water_works": (
        _import_wave3_utility_processor,
        "process_birmingham_water_works_batch",
    ),
    "city_of_mcminnville_water_sewer_dept": (
        _import_wave3_utility_processor,
        "process_city_of_mcminnville_water_sewer_dept_batch",
    ),
    "city_of_chattanooga_wastewater_department": (
        _import_wave3_utility_processor,
        "process_city_of_chattanooga_wastewater_department_batch",
    ),
    "city_of_martin": (
        _import_wave3_utility_processor,
        "process_city_of_martin_batch",
    ),
    "city_of_union_city": (
        _import_wave3_utility_processor,
        "process_city_of_union_city_batch",
    ),
    "guardian_water_power": (
        _import_wave3_utility_processor,
        "process_guardian_water_power_batch",
    ),
    "hopkinsville_electric_system": (
        _import_wave3_utility_processor,
        "process_hopkinsville_electric_system_batch",
    ),
    "cumberland_emc": (
        _import_wave3_utility_processor,
        "process_cumberland_emc_batch",
    ),
    "pleasant_view_utility_district": (
        _import_wave3_utility_processor,
        "process_pleasant_view_utility_district_batch",
    ),
}


# ---------------------------------------------------------------------------
# Per-batch processing entrypoint
# ---------------------------------------------------------------------------
def process_batch(
    batch_id: str,
    *,
    dry_run: bool = False,
    rules_override_paths: dict[str, "Path"] | None = None,
    only_filename: str | None = None,
    finalize_progress: bool = True,
    forced_vendor_key: str | None = None,
    route_filename: str | None = None,
    route_page: int | None = None,
) -> dict[str, Any]:
    """Run the appropriate vendor processor over every file in the batch.

    Phase 2A — adds two opt-in flags for the Vendor Rules Studio's
    "Test against batch" feature:

      * ``dry_run=True``   — propagated into ``run_context["dry_run"]``.
        Vendor processors that honour the flag skip Dropbox uploads and
        ResMan workbook writes; everything else runs identically. The
        webapp's preview cache (``_webapp_result.json``) is the caller's
        responsibility — this function never writes it (the API layer
        does, and skips it for dry-run calls).
      * ``rules_override_paths``  — ``{vendor_key: Path}`` mapping. When a
        vendor has an entry, that path is used as ``config_path`` instead
        of ``config/vendors/<vendor_key>.yaml``. The processor still calls
        ``yaml.safe_load()`` exactly the way the CLI does, so the rule
        loader is unchanged. Caller is responsible for cleaning up the
        temp files.

    Returns a dict shaped roughly like:
        {
            "batch_id": ...,
            "summary": {...},
            "by_vendor": {
                "richmond_utilities": ProcessBatchResult-as-dict,
            },
            "unsupported_files": [...]
        }

    Side-effect: writes a per-batch progress snapshot at
    `webapp_data/batches/<id>/progress.json` so the frontend can poll
    `GET /api/batches/<id>/progress` while the batch runs. (The progress
    file is written even on dry-run calls; nothing else on disk changes.)
    """
    rules_override_paths = rules_override_paths or {}
    # Private Deterministic Builder adapter. This can never alter production
    # routing: it is accepted only for an isolated dry-run and only for a
    # processor already present in the canonical registry.
    if forced_vendor_key:
        if not dry_run:
            raise ValueError("forced_vendor_key is permitted only for dry-run previews.")
        if forced_vendor_key not in _PROCESSOR_LOADERS:
            raise ValueError("forced_vendor_key is not a registered deterministic processor.")
    # Phase PERF-1 — optional perf timer. Safe import; no-op when the
    # module is missing or PERF_TIMER_DISABLED=1.
    try:
        from . import perf_timer as _perf  # type: ignore
    except Exception:  # pragma: no cover
        _perf = None  # type: ignore
    _t0_batch = time.perf_counter()
    bdir = batch_store.get_batch_dir(batch_id)
    in_dir = batch_store.get_input_dir(batch_id)
    processed_dir = batch_store.get_processed_dir(batch_id)
    files = batch_store.list_files_in_batch(batch_id)
    # Phase 2M — single-file processing. When ``only_filename`` is set,
    # narrow the file list to just that one file before vendor grouping
    # so the rest of the pipeline runs verbatim. Raises if the requested
    # filename isn't part of the batch (operator typo / stale UI).
    if only_filename:
        narrowed = [f for f in files if f.name == only_filename]
        if not narrowed:
            raise FileNotFoundError(
                f"File '{only_filename}' is not in batch '{batch_id}'.",
            )
        files = narrowed

    # Phase 1H: pull batch-level mode + AI policy + region hints from disk.
    batch_meta = _read_batch_metadata(batch_id)
    document_mode = batch_meta.get("document_mode") or "auto_detect"
    ai_fallback_policy = batch_meta.get("ai_fallback_policy") or "only_low_confidence"
    ai_fallback_enabled_meta = bool(batch_meta.get("ai_fallback_enabled", True))
    region_hints = _read_region_hints(batch_id)

    # AI fallback service — disabled by default. The service is the
    # single gate that decides whether any provider call ever fires.
    ai_service = None
    try:
        from .ai_fallback import get_service
        ai_service = get_service()
    except Exception:
        ai_service = None

    # Progress tracker — defensive import. CLI doesn't use this; the
    # webapp always wires it up.
    progress_path = bdir / "progress.json"
    tracker = None
    progress_callback = None
    try:
        from utils.progress_tracker import ProgressTracker, make_callback
        tracker = ProgressTracker(progress_path, batch_id=batch_id)
        tracker.declare_stages(_DEFAULT_STAGES)
        tracker.update(status="processing", files_total=len(files),
                       current_step="Detecting vendors…", percent=1.0)
        # Phase 1N — register the tracker so the /cancel endpoint can
        # flag it without round-tripping through the filesystem.
        try:
            from . import cancel_registry
            cancel_registry.register(batch_id, tracker)
        except Exception:
            pass
        # Upload stage was actually completed before we got here (the
        # files are already on disk), so flip it green right away.
        tracker.complete_stage("upload", detail=f"{len(files)} file(s)")
        tracker.start_stage("vendor_detect", detail=f"{len(files)} file(s)")
        raw_progress_callback = make_callback(tracker)
        if raw_progress_callback is not None and not finalize_progress:
            def progress_callback(**fields: Any) -> None:
                if fields.get("status") == "completed":
                    try:
                        percent = float(fields.get("percent") or 98.0)
                    except (TypeError, ValueError):
                        percent = 98.0
                    fields = {
                        **fields,
                        "status": "processing",
                        "percent": min(percent, 98.0),
                        "current_step": fields.get("current_step") or "Finalizing processor output...",
                    }
                raw_progress_callback(**fields)
        else:
            progress_callback = raw_progress_callback
    except Exception:
        tracker = None
        progress_callback = None

    # Phase 1N — cooperative cancellation hook. Vendor processors that
    # accept `should_cancel_callback` poll this between files / pages.
    # CLI runs (no tracker) get a no-op that always returns False so
    # behaviour is identical to today.
    def should_cancel() -> bool:
        return tracker is not None and tracker.is_cancel_requested()

    # Group files by detected vendor and AI authorization.  The second key is
    # essential: a batch may contain document/page overrides, and a paid AI
    # fallback authorized for one document must never leak into another.
    grouped: dict[tuple[str, bool], list[Path]] = {}
    detection: dict[str, dict] = {}
    route_decisions: dict[str, dict[str, Any]] = {}
    route_blocked: list[dict[str, Any]] = []
    # Phase PERF-1 — vendor detection is the first per-file hot spot;
    # measure it so the audit can spot regressions on large batches.
    _detect_cm = (
        _perf.perf_step("vendor.detect_all", batch_id=batch_id,
                        meta={"n_files": len(files)})
        if _perf is not None else None
    )
    if _detect_cm is not None:
        _detect_cm.__enter__()
    try:
        # Vendor detection is a routing decision, not an extraction pass.
        # Digital text and filename evidence stay available, while scanned
        # PDFs never run Tesseract here. Unknown scans go to the universal
        # AI/Vision path, which reads the page once and extracts the vendor
        # together with the invoice fields.
        with fast_detection_context():
            for f in files:
                det = ({
                    "vendor_key": forced_vendor_key,
                    "confidence": 1.0,
                    "source": "private_builder_dry_run_override",
                } if forced_vendor_key else detect_vendor_for_file(f))
                # Preserve deterministic processors for scanned utilities:
                # if cheap filename/text-layer routing is inconclusive, read
                # the scan once and apply only strong vendor keywords. The
                # ingestion OCR is file-hash cached, so the selected processor
                # or universal AI path reuses it instead of paying twice.
                if (
                    not forced_vendor_key
                    and det.get("vendor_key") == "unknown"
                    and f.suffix.lower() in {".pdf", ".png", ".jpg", ".jpeg", ".tif", ".tiff"}
                ):
                    try:
                        from . import document_ingestion

                        candidate = document_ingestion.ingest_document(f, max_pages=2)
                        strong = detect_vendor_from_text(f, candidate.document_text or "")
                        if strong is not None:
                            det = strong
                    except Exception:
                        pass
                policy_filename = (
                    route_filename
                    if only_filename and route_filename and f.name == only_filename
                    else f.name
                )
                policy_page = route_page if policy_filename == route_filename else None
                from .processing_route_gate import (
                    EffectiveProcessingRoute,
                    decide_processing_route,
                )
                from .processing_route_policy import resolve_requested_mode

                requested = resolve_requested_mode(
                    batch_id,
                    filename=policy_filename,
                    page=policy_page,
                )
                vendor_key = str(det.get("vendor_key") or "unknown")
                registered = vendor_key in _PROCESSOR_LOADERS
                processor_id = (
                    f"{vendor_key}.{_PROCESSOR_LOADERS[vendor_key][1]}"
                    if registered
                    else None
                )
                route = decide_processing_route(
                    requested,
                    vendor_key=vendor_key,
                    deterministic_available=registered,
                    processor_id=processor_id,
                )
                route_payload = route.model_dump(mode="json")
                route_decisions[policy_filename] = route_payload
                det = {**det, "route_decision": route_payload}
                detection[f.name] = det
                if route.effective_route == EffectiveProcessingRoute.BLOCKED:
                    route_blocked.append({
                        "filename": policy_filename,
                        "vendor_key": vendor_key,
                        "reason": route.reason_code,
                        "message": (
                            "Processing is locked to a deterministic route, but no "
                            "registered deterministic processor was available. AI was not called."
                        ),
                    })
                    continue
                grouped.setdefault(
                    (vendor_key, route.ai_fallback_authorized),
                    [],
                ).append(f)
    finally:
        if _detect_cm is not None:
            _detect_cm.__exit__(None, None, None)

    if tracker is not None:
        tracker.complete_stage(
            "vendor_detect",
            detail=f"{len(grouped)} vendor(s); document_mode={document_mode}",
        )

    by_vendor: dict[str, dict[str, Any]] = {}
    unsupported: list[dict] = list(route_blocked)
    overall_invoices: list[dict] = []
    overall_review: list[dict] = []

    if tracker is not None:
        tracker.update(current_step=f"Routing files to {len(grouped)} vendor(s)…", percent=3.0)

    for (vendor_key, route_ai_fallback_authorized), vfiles in grouped.items():
        # Phase 1N — bail before each vendor if cancellation was requested
        # while a previous vendor was running.
        if should_cancel():
            break
        if vendor_key not in _PROCESSOR_LOADERS:
            try:
                from . import ai_invoice_processor
                if ai_invoice_processor.should_route_to_ai(
                    vendor_key,
                    detection.get(vfiles[0].name) if vfiles else None,
                ):
                    ai_payload = ai_invoice_processor.process_ai_vendor_files(
                        batch_id=batch_id,
                        vendor_key=vendor_key,
                        files=vfiles,
                        detection=detection,
                        tracker=tracker,
                        should_cancel=should_cancel,
                        dry_run=dry_run,
                    )
                    ai_key = ai_payload.get("vendor_key") or ai_invoice_processor.AI_VENDOR_KEY
                    existing = by_vendor.get(ai_key)
                    if existing:
                        existing.setdefault("invoices", []).extend(
                            ai_payload.get("invoices") or []
                        )
                        existing.setdefault("manual_review_rows", []).extend(
                            ai_payload.get("manual_review_rows") or []
                        )
                        existing.setdefault("unsupported_files", []).extend(
                            ai_payload.get("unsupported_files") or []
                        )
                        old_summary = dict(existing.get("summary") or {})
                        new_summary = dict(ai_payload.get("summary") or {})
                        for key in (
                            "files_total",
                            "files_unique",
                            "files_deduplicated",
                            "files_processed",
                            "files_unique_processed",
                            "files_unsupported",
                            "processing_failures",
                            "invoices_produced",
                            "rows_total",
                            "line_items",
                            "manual_review_total",
                            "invoices_flagged_for_review",
                        ):
                            old_summary[key] = int(old_summary.get(key) or 0) + int(
                                new_summary.get(key) or 0
                            )
                        old_summary["processing_mode"] = "ai_assisted"
                        existing["summary"] = old_summary
                    else:
                        by_vendor[ai_key] = ai_payload
                    unsupported.extend(ai_payload.get("unsupported_files") or [])
                    overall_invoices.extend(ai_payload.get("invoices") or [])
                    overall_review.extend(ai_payload.get("manual_review_rows") or [])
                    continue
            except Exception:
                _LOG.exception("AI invoice routing failed for %s", vendor_key)

            for f in vfiles:
                unsupported.append({
                    "filename": f.name,
                    "vendor_key": vendor_key,
                    "detection": detection.get(f.name),
                    "reason": "no_processor_for_vendor_in_phase_1",
                })
            continue

        # Per-vendor working dirs inside the batch.
        vendor_in = bdir / "input" / vendor_key
        vendor_out = processed_dir / vendor_key
        vendor_in.mkdir(parents=True, exist_ok=True)
        vendor_out.mkdir(parents=True, exist_ok=True)
        # Stage files into the per-vendor input folder so the processor sees
        # only the files for THIS run. The staging folder is persistent
        # under webapp_data; if we leave older staged PDFs in place, a
        # single-file run will accidentally process the whole previous
        # vendor batch. We remove only staged files inside this sandbox,
        # never the operator's original uploads.
        current_names = {f.name for f in vfiles}
        for staged in vendor_in.iterdir():
            if staged.is_file() and staged.name not in current_names:
                try:
                    staged.unlink()
                except OSError:
                    pass
        for f in vfiles:
            target = vendor_in / f.name
            try:
                if target.exists():
                    target.unlink()
                shutil.copy2(f, target)
            except Exception as e:
                unsupported.append({
                    "filename": f.name,
                    "vendor_key": vendor_key,
                    "reason": f"failed_to_stage_input:{type(e).__name__}",
                })

        # Call the processor via the registry tuple (loader, entrypoint_name).
        loader, entrypoint_name = _PROCESSOR_LOADERS[vendor_key]
        mod = loader()
        process_func = getattr(mod, entrypoint_name, None)
        if process_func is None:
            unsupported.extend([
                {"filename": f.name, "vendor_key": vendor_key,
                 "reason": f"vendor_processor_function_missing:{entrypoint_name}"}
                for f in vfiles
            ])
            continue

        if tracker is not None:
            tracker.update(current_step=f"Processing {len(vfiles)} {vendor_key} file(s)…",
                           percent=5.0)

        # Phase 2A — Rules Studio impact preview: caller can pass an
        # alternate YAML path per vendor so we test draft rules without
        # touching the on-disk file.
        config_path = rules_override_paths.get(vendor_key) or (
            VENDORS_DIR / f"{vendor_key}.yaml"
        )
        # Pass progress_callback when the processor accepts it. Older
        # processors that don't will get a TypeError at call time — we
        # introspect the signature to stay backwards compatible.
        import inspect
        # Phase 1H — enrich `run_context` with batch-level mode, AI
        # policy, region hints, and a reference to the AI service.
        # Vendor processors that don't read these keys ignore them; the
        # CLI sets `run_context=None` and behaves identically to today.
        run_context: dict[str, Any] = {
            "timestamp": datetime.now().strftime("%Y%m%d_%H%M%S"),
            "source": "webapp",
            "batch_id": batch_id,
            "document_mode": document_mode,
            "ai_fallback_enabled": (
                ai_fallback_enabled_meta
                and route_ai_fallback_authorized
                and ai_service is not None
                and ai_service.is_enabled()
            ),
            "ai_fallback_policy": ai_fallback_policy,
            "ai_fallback_service": ai_service,
            "region_hints": [
                r for r in region_hints
                if (r.get("file_id") or "") in {f.name for f in vfiles}
            ],
            # Phase 2A — vendor processors that honour this skip Dropbox
            # uploads and Excel writes. Default False keeps CLI + normal
            # webapp runs identical to today.
            "dry_run": dry_run,
            "processing_route_decisions": [
                route_decisions.get(
                    route_filename
                    if only_filename and route_filename and f.name == only_filename
                    else f.name,
                    {},
                )
                for f in vfiles
            ],
        }
        kwargs: dict[str, Any] = {
            "input_folder": vendor_in,
            "output_folder": vendor_out,
            "template_path": RESMAN_TEMPLATE,
            "config_path": config_path,
            "run_context": run_context,
        }
        try:
            sig = inspect.signature(process_func)
            if "progress_callback" in sig.parameters and progress_callback is not None:
                kwargs["progress_callback"] = progress_callback
            # Phase 1N — pass should_cancel_callback when supported.
            if "should_cancel_callback" in sig.parameters:
                kwargs["should_cancel_callback"] = should_cancel
        except (TypeError, ValueError):
            pass

        # Phase 1N — also expose the cancel callable inside run_context
        # so processors that read `run_context` directly can poll it.
        run_context["should_cancel"] = should_cancel

        if tracker is not None:
            # The vendor processor is going to read text, optionally OCR,
            # apply YAML rules, match addresses, etc. — without a
            # processor-level hook we can't drive each individual stage,
            # but we can mark the broader stages as we enter / leave the
            # vendor call.
            for k in ("read_pdf", "ocr", "yaml_rules", "address_match",
                      "unit_match", "gl_evidence"):
                tracker.start_stage(k, detail=f"{vendor_key}: {len(vfiles)} file(s)")
        # Phase PERF-1 — measure each vendor's total wall-clock.
        _proc_cm = (
            _perf.perf_step(f"processor.{vendor_key}", batch_id=batch_id,
                            meta={"n_files": len(vfiles)})
            if _perf is not None else None
        )
        if _proc_cm is not None:
            _proc_cm.__enter__()
        try:
            result = process_func(**kwargs)
        finally:
            if _proc_cm is not None:
                _proc_cm.__exit__(None, None, None)
        if tracker is not None:
            for k in ("read_pdf", "ocr", "yaml_rules", "address_match",
                      "unit_match", "gl_evidence"):
                tracker.complete_stage(k)
            if run_context["ai_fallback_enabled"]:
                tracker.start_stage("ai_fallback",
                                    detail=f"policy={ai_fallback_policy}")
                tracker.complete_stage("ai_fallback")
            else:
                tracker.skip_stage(
                    "ai_fallback",
                    detail="disabled or not configured",
                )
            tracker.start_stage("reconcile")
            tracker.complete_stage("reconcile")
            tracker.start_stage("split_pdf")
            tracker.complete_stage("split_pdf")
            tracker.start_stage("dropbox")
            tracker.complete_stage("dropbox")
            tracker.start_stage("template")
            tracker.complete_stage("template")
        # ProcessBatchResult is a dataclass — convert to plain dict.
        from dataclasses import asdict
        result_dict = asdict(result)
        if vendor_key in by_vendor:
            by_vendor[vendor_key] = _merge_vendor_payload(
                by_vendor[vendor_key], result_dict,
            )
        else:
            by_vendor[vendor_key] = result_dict

        # A deterministic processor must never consume a file silently. A
        # false-positive detector or changed layout can yield zero invoices
        # without raising. Route that exact file set through the universal AI
        # path and retain the empty deterministic payload for auditability.
        if _needs_zero_output_ai_fallback(result, vfiles):
            if not route_ai_fallback_authorized:
                for f in vfiles:
                    policy_filename = (
                        route_filename
                        if only_filename and route_filename and f.name == only_filename
                        else f.name
                    )
                    detection[f.name]["fallback_processing_mode"] = "blocked_by_route_policy"
                    unsupported.append({
                        "filename": policy_filename,
                        "vendor_key": vendor_key,
                        "reason": "deterministic_zero_output_ai_not_authorized",
                        "message": (
                            "The deterministic processor produced no invoice disposition. "
                            "Cost-safe routing prohibited a silent AI fallback; manual review "
                            "or an explicit AI-fallback authorization is required."
                        ),
                    })
                continue
            from . import ai_invoice_processor

            fallback_detection = {
                f.name: {
                    "vendor_key": "unknown",
                    "confidence": 0.0,
                    "reason": f"deterministic_zero_output:{vendor_key}",
                    "supported_in_phase_1": False,
                    "processing_mode": "ai_assisted",
                }
                for f in vfiles
            }
            for f in vfiles:
                detection[f.name]["initial_vendor_key"] = vendor_key
                detection[f.name]["fallback_processing_mode"] = "ai_assisted_zero_output"
            ai_payload = ai_invoice_processor.process_ai_vendor_files(
                batch_id=batch_id,
                vendor_key="unknown",
                files=vfiles,
                detection=fallback_detection,
                tracker=tracker,
                should_cancel=should_cancel,
                dry_run=dry_run,
            )
            ai_key = ai_payload.get("vendor_key") or ai_invoice_processor.AI_VENDOR_KEY
            existing_ai = by_vendor.get(ai_key)
            if existing_ai:
                existing_ai.setdefault("invoices", []).extend(ai_payload.get("invoices") or [])
                existing_ai.setdefault("manual_review_rows", []).extend(
                    ai_payload.get("manual_review_rows") or []
                )
                existing_ai.setdefault("unsupported_files", []).extend(
                    ai_payload.get("unsupported_files") or []
                )
                old_summary = dict(existing_ai.get("summary") or {})
                new_summary = dict(ai_payload.get("summary") or {})
                for key in (
                    "files_total", "files_unique", "files_deduplicated",
                    "files_processed", "files_unique_processed", "files_unsupported",
                    "processing_failures",
                    "invoices_produced", "rows_total", "line_items",
                    "manual_review_total", "invoices_flagged_for_review",
                ):
                    old_summary[key] = int(old_summary.get(key) or 0) + int(
                        new_summary.get(key) or 0
                    )
                old_summary["processing_mode"] = "ai_assisted"
                existing_ai["summary"] = old_summary
            else:
                by_vendor[ai_key] = ai_payload
            fallback_invoices = list(ai_payload.get("invoices") or [])
            fallback_review = list(ai_payload.get("manual_review_rows") or [])
            fallback_unsupported = list(ai_payload.get("unsupported_files") or [])
            if not (fallback_invoices or fallback_review or fallback_unsupported):
                fallback_unsupported = [{
                    "filename": f.name,
                    "vendor_key": vendor_key,
                    "reason": "deterministic_and_ai_zero_output",
                    "message": "No invoice rows were produced; manual review is required.",
                } for f in vfiles]
            overall_invoices.extend(fallback_invoices)
            overall_review.extend(fallback_review)
            unsupported.extend(fallback_unsupported)
            continue
        overall_invoices.extend(result.invoices)
        overall_review.extend(result.manual_review_rows)

    # Top-level summary
    ai_deduplicated_files = sum(
        int((payload.get("summary") or {}).get("files_deduplicated") or 0)
        for payload in by_vendor.values()
        if (payload.get("summary") or {}).get("processing_mode") == "ai_assisted"
    )
    failed_source_files = {
        str(item.get("filename") or "").strip()
        for item in unsupported
        if isinstance(item, dict) and str(item.get("filename") or "").strip()
    }
    summary = {
        "files_total": len(files),
        "files_unique": max(0, len(files) - ai_deduplicated_files),
        "files_deduplicated": ai_deduplicated_files,
        "files_supported": _supported_file_count(files, unsupported),
        "files_unsupported": len(failed_source_files),
        "processing_failures": len(unsupported),
        "invoices_total": len(overall_invoices),
        "manual_review_total": len(overall_review),
    }

    if tracker is not None:
        if should_cancel():
            # Phase 1N — finalise as cancelled rather than completed.
            tracker.cancelled(
                files_total=len(files),
                files_done=sum(
                    int((payload.get("summary") or {}).get("files_processed") or 0)
                    for payload in by_vendor.values()
                ),
                invoices_created=len(overall_invoices),
                rows_created=sum(
                    len(inv.get("rows", [])) for inv in overall_invoices
                ),
                warnings_count=len(overall_review),
            )
            summary["cancelled"] = True
        else:
            if not by_vendor:
                for k in (
                    "read_pdf",
                    "ocr",
                    "yaml_rules",
                    "address_match",
                    "unit_match",
                    "gl_evidence",
                    "ai_fallback",
                    "reconcile",
                    "split_pdf",
                    "dropbox",
                    "template",
                ):
                    tracker.skip_stage(k, detail="No supported files in batch")
            tracker.start_stage("ready")
            tracker.complete_stage(
                "ready",
                detail=f"{len(overall_invoices)} invoice(s), {len(overall_review)} flagged",
            )
            summary_fields = {
                "files_total": len(files),
                "files_done": len(files),
                "invoices_created": len(overall_invoices),
                "rows_created": sum(len(inv.get("rows", [])) for inv in overall_invoices),
                "warnings_count": len(overall_review),
            }
            if finalize_progress:
                tracker.complete(**summary_fields)
            else:
                tracker.update(
                    status="processing",
                    percent=99.0,
                    current_step="Finalizing preview...",
                    **summary_fields,
                )
        # Phase 1N — release the registry entry whether we completed
        # normally or cancelled.
        try:
            from . import cancel_registry
            cancel_registry.unregister(batch_id)
        except Exception:
            pass

    # Phase PERF-1 — record the batch total and flush timings to disk.
    if _perf is not None:
        try:
            _total_ms = (time.perf_counter() - _t0_batch) * 1000.0
            _perf.record(batch_id, "batch.total", _total_ms,
                         meta={"files": len(files),
                                "vendors": len(grouped),
                                "cancelled": bool(should_cancel())})
            _perf.flush_to_disk(batch_id, bdir / "audit")
        except Exception:  # pragma: no cover
            _LOG.exception("perf_timer flush failed for %s", batch_id)

    return {
        "batch_id": batch_id,
        "summary": summary,
        "by_vendor": by_vendor,
        "detection": detection,
        "processing_routes": route_decisions,
        "unsupported_files": unsupported,
        "all_invoices": overall_invoices,
        "all_manual_review": overall_review,
    }


def _needs_zero_output_ai_fallback(result: Any, files: list[Path]) -> bool:
    """Return True when a processor consumed files without any disposition."""
    return bool(files) and not (
        list(getattr(result, "invoices", None) or [])
        or list(getattr(result, "manual_review_rows", None) or [])
        or list(getattr(result, "errors", None) or [])
    )


def _merge_vendor_payload(
    existing: dict[str, Any], incoming: dict[str, Any],
) -> dict[str, Any]:
    """Merge repeated policy-partition runs for one deterministic vendor."""

    merged = dict(existing)
    list_keys = {
        "invoices", "manual_review_rows", "unsupported_files", "errors",
        "processed_files", "output_files",
    }
    for key, value in incoming.items():
        if key == "summary" or value is None:
            continue
        if key in list_keys and isinstance(value, list):
            merged[key] = [*(merged.get(key) or []), *value]
        elif key not in merged or merged.get(key) in (None, "", [], {}):
            merged[key] = value
    old_summary = dict(existing.get("summary") or {})
    new_summary = dict(incoming.get("summary") or {})
    summary = dict(old_summary)
    for key, value in new_summary.items():
        if (
            isinstance(value, (int, float))
            and not isinstance(value, bool)
            and isinstance(old_summary.get(key, 0), (int, float))
            and not isinstance(old_summary.get(key, 0), bool)
        ):
            summary[key] = old_summary.get(key, 0) + value
        elif key not in summary:
            summary[key] = value
        elif summary[key] != value and key == "processing_mode":
            summary[key] = "mixed_policy_partitions"
    merged["summary"] = summary
    return merged


# ---------------------------------------------------------------------------
# Helper: write edited preview rows into a fresh copy of Output/Template.xlsx
# ---------------------------------------------------------------------------
# Keys the frontend may include but which aren't real ResMan columns. Anything
# starting with "_" is also stripped.
_NON_TEMPLATE_KEYS = {"_meta", "_edited", "_row_index"}

_LOCAL_DOCUMENT_URL_HOSTS = {"localhost", "127.0.0.1", "::1"}


def _is_dropbox_document_url(value: Any) -> bool:
    text = str(value or "").strip()
    if not text:
        return False
    try:
        host = (urlparse(text).netloc or "").lower()
    except Exception:
        return False
    return host == "dropbox.com" or host.endswith(".dropbox.com") or host.endswith(".dropboxusercontent.com")


def _is_local_document_url(value: Any) -> bool:
    text = str(value or "").strip()
    if not text:
        return False
    if text.startswith("/api/batches/"):
        return True
    try:
        parsed = urlparse(text)
    except Exception:
        return False
    host = (parsed.hostname or "").lower()
    return host in _LOCAL_DOCUMENT_URL_HOSTS


def _local_document_url(batch_id: str, source_file: str) -> str:
    base = "http://localhost:8001"
    try:
        from . import row_normalizer

        base = row_normalizer._webapp_base_url()
    except Exception:
        pass
    return (
        base.rstrip("/")
        + f"/api/batches/{quote(batch_id, safe='')}/files/{quote(Path(source_file).name, safe='')}/content"
    )


def _allow_local_document_url_export() -> bool:
    """Development escape hatch for offline demos only.

    Production exports must not silently write localhost document links into
    ResMan templates. Operators can still enable this explicitly while testing
    a disconnected dev box.
    """

    value = os.environ.get("ALLOW_LOCAL_DOCUMENT_URL_EXPORT", "")
    return value.strip().lower() in {"1", "true", "yes", "on"}


def _write_edited_rows_to_template(template_path: Path, dest: Path,
                                   rows: list[dict[str, Any]]) -> int:
    """Copy the official ResMan template and write `rows` into it. Each
    dict in `rows` is matched to template columns by exact header name;
    missing columns are left blank; extra keys (like `_meta`) are ignored.
    Returns the number of rows actually written."""
    dest.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template_path, dest)
    wb = openpyxl.load_workbook(dest)
    ws = wb["Sheet 1"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    header_to_col: dict[str, int] = {}
    for i, h in enumerate(headers, start=1):
        if h is None:
            continue
        key = str(h).strip()
        if key:
            header_to_col[key] = i

    # Phase 2L — Date columns are written as real Excel dates so
    # ResMan accepts them as date cells, not text.
    try:
        from utils.excel_helpers import set_cell as _set_cell
    except Exception:  # pragma: no cover
        _set_cell = None  # type: ignore

    written = 0
    for r_idx, row in enumerate(rows, start=2):
        for key, value in row.items():
            if key in _NON_TEMPLATE_KEYS or key.startswith("_"):
                continue
            col = header_to_col.get(key)
            if not col:
                continue
            cell = ws.cell(row=r_idx, column=col)
            if _set_cell is not None:
                _set_cell(cell, _coerce_cell_value(value), key)
            else:
                cell.value = _coerce_cell_value(value)
        written += 1
    wb.save(dest)
    return written


def _validate_resman_export_required_rows(rows: list[dict[str, Any]]) -> None:
    """Block exports that would create unusable ResMan rows.

    Required columns are operator-configurable from the Formats workspace.
    The preview/review UI can show missing values, but the actual exported
    template must not be emitted until configured required fields are resolved.
    """
    required_columns = [
        col for col in get_template_rules().get("required_columns", [])
        if str(col or "").strip()
    ]
    missing: list[str] = []
    for idx, row in enumerate(rows, start=1):
        amount = row.get("Amount")
        try:
            payable = abs(float(str(amount).replace(",", "") or 0)) > 0.009
        except (TypeError, ValueError):
            payable = bool(str(amount or "").strip())
        if not payable:
            continue
        row_missing = [
            col for col in required_columns
            if not _has_required_export_value(row.get(col))
        ]
        if row_missing:
            invoice = str(row.get("Invoice Number") or "").strip() or f"row {idx}"
            missing.append(f"{invoice} line {row.get('Line Item Number') or idx}: {', '.join(row_missing)}")
    if missing:
        shown = "; ".join(missing[:6])
        more = "" if len(missing) <= 6 else f"; +{len(missing) - 6} more"
        raise ValueError(
            "Required template fields are missing before export. Resolve: "
            f"{shown}{more}"
        )


def _has_required_export_value(value: Any) -> bool:
    """Return True when an export-required cell has an intentional value.

    Boolean ``False`` is a valid ResMan value for "Is Replacement Reserve";
    the previous ``row.get(col) or ""`` check treated it as blank and blocked
    otherwise valid exports.
    """

    if value is None:
        return False
    if isinstance(value, bool):
        return True
    return bool(str(value).strip())


def _coerce_cell_value(value: Any) -> Any:
    """Light coercion so the resulting xlsx renders nicely:
    - strings that look like numbers stay as strings (we don't want to
      autoformat invoice numbers)
    - ``None`` and empty string render as blank cells
    - everything else passes through
    """
    if value is None:
        return None
    if isinstance(value, str):
        return value if value != "" else None
    return value


# ---------------------------------------------------------------------------
# Final export — copy the per-vendor ResMan workbook to the batch export
# folder, OR, if `edited_rows` is supplied by the frontend, write a fresh
# workbook from the official template using those edited values.
# ---------------------------------------------------------------------------
def _normalize_edited_rows_for_export(batch_id: str, edited_rows: list[dict[str, Any]]) -> None:
    """Reapply final row cleanup before edited-row export validation."""

    if not edited_rows:
        return
    try:
        from . import row_normalizer
    except Exception:  # pragma: no cover - export should still validate normally
        _LOG.debug("Row normalizer unavailable during edited export", exc_info=True)
        return

    source_by_invoice: dict[str, str] = {}
    try:
        cache = batch_store.get_processed_dir(batch_id) / "_webapp_result.json"
        if cache.is_file():
            cached = json.loads(cache.read_text(encoding="utf-8")) or {}
            for inv in list(cached.get("all_invoices") or []):
                debug_info = inv.get("debug_info") if isinstance(inv.get("debug_info"), dict) else {}
                src = str(inv.get("source_file") or debug_info.get("source_file") or "").strip()
                if not src:
                    continue
                for cached_row in list(inv.get("rows") or []):
                    inv_no = str((cached_row or {}).get("Invoice Number") or "").strip()
                    if inv_no and inv_no not in source_by_invoice:
                        source_by_invoice[inv_no] = src
    except Exception:
        _LOG.debug("Could not load processed cache for export source mapping", exc_info=True)

    single_source_file = ""
    try:
        files = [p.name for p in batch_store.get_input_dir(batch_id).rglob("*") if p.is_file()]
        if len(files) == 1:
            single_source_file = files[0]
    except Exception:
        single_source_file = ""

    grouped: dict[str, list[dict[str, Any]]] = {}
    ungrouped: list[dict[str, Any]] = []
    for row in edited_rows:
        if not isinstance(row, dict):
            continue
        meta = row.get("_meta") if isinstance(row.get("_meta"), dict) else {}
        src = str(
            (meta or {}).get("source_file")
            or row.get("source_file")
            or row.get("Source File")
            or source_by_invoice.get(str(row.get("Invoice Number") or "").strip(), "")
            or single_source_file
            or ""
        ).strip()
        if src:
            grouped.setdefault(src, []).append(row)
        else:
            ungrouped.append(row)

    for src, rows in grouped.items():
        row_normalizer.normalize_rows(rows, batch_id=batch_id, source_file=src)
    if ungrouped:
        row_normalizer.normalize_rows(ungrouped, batch_id=batch_id)


def _source_mapping_from_preview_cache(batch_id: str) -> dict[str, str]:
    """Map invoice numbers to source files using the persisted preview cache."""

    cache = batch_store.get_processed_dir(batch_id) / "_webapp_result.json"
    if not cache.is_file():
        return {}
    try:
        cached = json.loads(cache.read_text(encoding="utf-8")) or {}
    except Exception:
        _LOG.debug("Could not read preview cache for export source mapping", exc_info=True)
        return {}

    source_by_invoice: dict[str, str] = {}
    for inv in list(cached.get("all_invoices") or []):
        if not isinstance(inv, dict):
            continue
        debug_info = inv.get("debug_info") if isinstance(inv.get("debug_info"), dict) else {}
        src = str(inv.get("source_file") or debug_info.get("source_file") or "").strip()
        if not src:
            continue
        for cached_row in list(inv.get("rows") or []):
            if not isinstance(cached_row, dict):
                continue
            inv_no = str(cached_row.get("Invoice Number") or "").strip()
            if inv_no and inv_no not in source_by_invoice:
                source_by_invoice[inv_no] = src
    return source_by_invoice


def _single_input_source_file(batch_id: str) -> str:
    try:
        files = [p.name for p in batch_store.get_input_dir(batch_id).rglob("*") if p.is_file()]
    except Exception:
        return ""
    return files[0] if len(files) == 1 else ""


def _row_source_file_for_export(
    batch_id: str,
    row: dict[str, Any],
    *,
    source_by_invoice: dict[str, str],
    single_source_file: str,
) -> str:
    meta = row.get("_meta") if isinstance(row.get("_meta"), dict) else {}
    inv_no = str(row.get("Invoice Number") or "").strip()
    return str(
        (meta or {}).get("source_file")
        or row.get("source_file")
        or row.get("Source File")
        or source_by_invoice.get(inv_no, "")
        or single_source_file
        or ""
    ).strip()


def _ensure_dropbox_document_urls_for_export(batch_id: str, rows: list[dict[str, Any]]) -> list[str]:
    """Replace local preview links with Dropbox links before writing Excel.

    Preview rows may carry a localhost ``Document Url`` so the operator can
    open the source document inside the webapp. The exported ResMan template
    must contain shareable Dropbox URLs, matching legacy batches like Zillow.
    """

    if not rows:
        return []

    try:
        from . import support_documents
    except Exception as exc:  # pragma: no cover
        _LOG.warning("Dropbox support-document service unavailable during export", exc_info=exc)
        message = "Dropbox support-document service is unavailable. Export was blocked to avoid localhost document links."
        if _allow_local_document_url_export():
            return [
                "Dropbox support-document service is unavailable. Export continued with local document links because "
                "ALLOW_LOCAL_DOCUMENT_URL_EXPORT is enabled."
            ]
        raise ValueError(message)

    source_by_invoice = _source_mapping_from_preview_cache(batch_id)
    single_source_file = _single_input_source_file(batch_id)
    uploads: dict[str, str] = {}
    failures: list[str] = []

    for row in rows:
        if not isinstance(row, dict):
            continue
        current = str(row.get("Document Url") or "").strip()
        if _is_dropbox_document_url(current):
            continue
        if current and not _is_local_document_url(current):
            # Preserve real external links supplied by a processor/operator.
            continue

        source_file = _row_source_file_for_export(
            batch_id,
            row,
            source_by_invoice=source_by_invoice,
            single_source_file=single_source_file,
        )
        if not source_file:
            failures.append("source file missing")
            continue
        if not current:
            row["Document Url"] = _local_document_url(batch_id, source_file)
            meta = row.setdefault("_meta", {})
            if isinstance(meta, dict):
                meta.setdefault("support_document_status", "local_webapp_link")
        if source_file not in uploads:
            link = support_documents.upload_source_document_to_dropbox(
                batch_id=batch_id,
                source_file=source_file,
                vendor_name=str(row.get("Vendor") or "Support Documents"),
                invoice_date=row.get("Invoice Date") or row.get("Accounting Date"),
                dry_run=False,
            )
            if not link.success:
                failures.append(
                    f"{source_file}: {link.review_message or link.status or 'Dropbox upload failed'}"
                )
                continue
            uploads[source_file] = link.url
        row["Document Url"] = uploads[source_file]
        meta = row.setdefault("_meta", {})
        if isinstance(meta, dict):
            meta["support_document_status"] = "dropbox_uploaded"

    unresolved_local = [
        str(row.get("Document Url") or "").strip()
        for row in rows
        if isinstance(row, dict)
        and (
            not str(row.get("Document Url") or "").strip()
            or _is_local_document_url(row.get("Document Url"))
        )
    ]
    if unresolved_local:
        detail = "; ".join(failures[:3]) if failures else "Dropbox upload did not return a shared URL."
        for row in rows:
            if not isinstance(row, dict):
                continue
            url = str(row.get("Document Url") or "").strip()
            if url and not _is_local_document_url(url):
                continue
            meta = row.setdefault("_meta", {})
            if isinstance(meta, dict):
                meta["support_document_status"] = "local_export_dropbox_unavailable"
                meta["support_document_warning"] = detail
        if not _allow_local_document_url_export():
            raise ValueError(
                "Dropbox could not create shared support-document links, so export was blocked to avoid "
                f"{len(unresolved_local)} localhost document link(s). "
                "Retry Export when Dropbox is reachable or review Dropbox credentials/network access. "
                + detail
            )
        return [
            "Dropbox could not create shared support-document links. "
            f"Export continued with {len(unresolved_local)} local document link(s). "
            "ALLOW_LOCAL_DOCUMENT_URL_EXPORT is enabled; re-export later to replace them with Dropbox links. "
            + detail
        ]
    return []


def _persist_export_document_urls_to_cache(batch_id: str, rows: list[dict[str, Any]]) -> None:
    """Best-effort: write exported Dropbox links back into preview cache."""

    if not rows:
        return
    source_by_invoice = _source_mapping_from_preview_cache(batch_id)
    single_source_file = _single_input_source_file(batch_id)
    url_by_source: dict[str, str] = {}
    for row in rows:
        if not isinstance(row, dict):
            continue
        url = str(row.get("Document Url") or "").strip()
        if not _is_dropbox_document_url(url):
            continue
        src = _row_source_file_for_export(
            batch_id,
            row,
            source_by_invoice=source_by_invoice,
            single_source_file=single_source_file,
        )
        if src:
            url_by_source[src] = url
    if not url_by_source:
        return

    cache = batch_store.get_processed_dir(batch_id) / "_webapp_result.json"
    if not cache.is_file():
        return
    try:
        cached = json.loads(cache.read_text(encoding="utf-8")) or {}
    except Exception:
        _LOG.debug("Could not read preview cache while persisting Dropbox links", exc_info=True)
        return

    changed = False
    for inv in list(cached.get("all_invoices") or []):
        if not isinstance(inv, dict):
            continue
        debug_info = inv.get("debug_info") if isinstance(inv.get("debug_info"), dict) else {}
        src = str(inv.get("source_file") or debug_info.get("source_file") or "").strip()
        url = url_by_source.get(src)
        if not url:
            continue
        for row in list(inv.get("rows") or []):
            if not isinstance(row, dict):
                continue
            if row.get("Document Url") != url:
                row["Document Url"] = url
                changed = True
            meta = row.setdefault("_meta", {})
            if isinstance(meta, dict):
                meta["support_document_status"] = "dropbox_uploaded"
    by_vendor = cached.get("by_vendor") if isinstance(cached.get("by_vendor"), dict) else {}
    for payload in by_vendor.values():
        for inv in list((payload or {}).get("invoices") or []):
            if not isinstance(inv, dict):
                continue
            debug_info = inv.get("debug_info") if isinstance(inv.get("debug_info"), dict) else {}
            src = str(inv.get("source_file") or debug_info.get("source_file") or "").strip()
            url = url_by_source.get(src)
            if not url:
                continue
            for row in list(inv.get("rows") or []):
                if not isinstance(row, dict):
                    continue
                if row.get("Document Url") != url:
                    row["Document Url"] = url
                    changed = True
                meta = row.setdefault("_meta", {})
                if isinstance(meta, dict):
                    meta["support_document_status"] = "dropbox_uploaded"
    if changed:
        try:
            cache.write_text(json.dumps(cached, default=str, indent=2), encoding="utf-8")
        except Exception:
            _LOG.debug("Could not persist Dropbox links to preview cache", exc_info=True)


def _document_url_updates_for_rows(
    batch_id: str,
    rows: list[dict[str, Any]],
) -> dict[str, dict[str, str]]:
    """Return Dropbox URL patches the frontend can apply to preview rows."""

    source_by_invoice = _source_mapping_from_preview_cache(batch_id)
    single_source_file = _single_input_source_file(batch_id)
    by_source_file: dict[str, str] = {}
    by_invoice_number: dict[str, str] = {}
    for row in rows:
        if not isinstance(row, dict):
            continue
        url = str(row.get("Document Url") or "").strip()
        if not _is_dropbox_document_url(url):
            continue
        src = _row_source_file_for_export(
            batch_id,
            row,
            source_by_invoice=source_by_invoice,
            single_source_file=single_source_file,
        )
        if src:
            by_source_file[src] = url
        inv_no = str(row.get("Invoice Number") or "").strip()
        if inv_no:
            by_invoice_number[inv_no] = url
    return {
        "by_source_file": by_source_file,
        "by_invoice_number": by_invoice_number,
    }


def _cached_preview_rows_for_export(batch_id: str) -> list[dict[str, Any]]:
    """Return rows from the web preview cache for batches without workbooks."""

    cache = batch_store.get_processed_dir(batch_id) / "_webapp_result.json"
    if not cache.is_file():
        return []
    try:
        cached = json.loads(cache.read_text(encoding="utf-8")) or {}
    except Exception:
        _LOG.debug("Could not read preview cache for export", exc_info=True)
        return []

    rows: list[dict[str, Any]] = []
    for inv in list(cached.get("all_invoices") or []):
        if not isinstance(inv, dict):
            continue
        debug_info = inv.get("debug_info") if isinstance(inv.get("debug_info"), dict) else {}
        src = str(inv.get("source_file") or debug_info.get("source_file") or "").strip()
        validation = inv.get("validation_summary") if isinstance(inv.get("validation_summary"), dict) else {}
        invoice_total = validation.get("invoice_total") or inv.get("total_amount")
        for raw_row in list(inv.get("rows") or []):
            if not isinstance(raw_row, dict):
                continue
            row = dict(raw_row)
            if src:
                meta = row.get("_meta") if isinstance(row.get("_meta"), dict) else {}
                meta = dict(meta)
                meta.setdefault("source_file", src)
                row["_meta"] = meta
            meta = row.get("_meta") if isinstance(row.get("_meta"), dict) else {}
            meta = dict(meta)
            if "total_reconciliation_passed" in validation:
                meta["total_reconciliation_passed"] = validation.get("total_reconciliation_passed")
            provenance = meta.get("ai_provenance") if isinstance(meta.get("ai_provenance"), dict) else {}
            if invoice_total is not None:
                provenance = dict(provenance)
                provenance.setdefault("invoice_total", invoice_total)
                meta["ai_provenance"] = provenance
            row["_meta"] = meta
            rows.append(row)
    return rows


def cached_preview_rows_for_readiness(batch_id: str) -> list[dict[str, Any]]:
    """Public compatibility adapter used by readiness/API/export callers."""
    return _cached_preview_rows_for_export(batch_id)


def _authorize_export(batch_id: str, rows: list[dict[str, Any]]) -> dict[str, Any]:
    from . import accounting_readiness
    decision = accounting_readiness.evaluate_and_record(batch_id, rows)
    if not decision.export_allowed:
        codes = ", ".join(issue.code for issue in decision.blockers[:8])
        raise ValueError(
            f"Export blocked by accounting readiness {decision.snapshot_id}: {codes}"
        )
    return accounting_readiness.as_dict(decision)


def export_batch(batch_id: str, edited_rows: Optional[list[dict[str, Any]]] = None) -> dict[str, Any]:
    bdir = batch_store.get_batch_dir(batch_id)
    processed_dir = bdir / "processed"
    export_dir = batch_store.get_export_dir(batch_id)

    # ---- Path A: edited export (frontend sent the table state) ----
    if edited_rows is not None:
        _normalize_edited_rows_for_export(batch_id, edited_rows)
        document_url_warnings = _ensure_dropbox_document_urls_for_export(batch_id, edited_rows)
        _persist_export_document_urls_to_cache(batch_id, edited_rows)
        document_url_updates = _document_url_updates_for_rows(batch_id, edited_rows)
        from .validated_export_bridge import ExportAuthorizationError, ReadinessValidatedExporter
        try:
            readiness = ReadinessValidatedExporter().authorize(batch_id, edited_rows)
        except ExportAuthorizationError as exc:
            return {"batch_id": batch_id, "exported": [], "reason": exc.code, "blockers": exc.blockers}
        _validate_resman_export_required_rows(edited_rows)
        if not RESMAN_TEMPLATE.is_file():
            return {
                "batch_id": batch_id, "exported": [],
                "reason": "template_missing", "template_path": str(RESMAN_TEMPLATE),
            }
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        dest = export_dir / f"resman_import_edited_{ts}.xlsx"
        rows_written = _write_edited_rows_to_template(RESMAN_TEMPLATE, dest, edited_rows)
        return {
            "batch_id": batch_id,
            "exported": [{
                "vendor_key": "edited",
                "source_path": str(RESMAN_TEMPLATE),
                "export_path": str(dest),
                "filename": dest.name,
            }],
            "export_used_edited_rows": True,
            "edited_rows_count": len(edited_rows),
            "rows_written": rows_written,
            "document_url_updates": document_url_updates,
            "document_url_warnings": document_url_warnings,
            "accounting_readiness": readiness,
        }

    # ---- Path B: legacy export (copy latest per-vendor processed xlsx) ----
    if not processed_dir.is_dir():
        preview_rows = _cached_preview_rows_for_export(batch_id)
        if preview_rows:
            _normalize_edited_rows_for_export(batch_id, preview_rows)
            readiness = _authorize_export(batch_id, preview_rows)
            document_url_warnings = _ensure_dropbox_document_urls_for_export(batch_id, preview_rows)
            _persist_export_document_urls_to_cache(batch_id, preview_rows)
            document_url_updates = _document_url_updates_for_rows(batch_id, preview_rows)
            _validate_resman_export_required_rows(preview_rows)
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            dest = export_dir / f"resman_import_preview_{ts}.xlsx"
            rows_written = _write_edited_rows_to_template(RESMAN_TEMPLATE, dest, preview_rows)
            return {
                "batch_id": batch_id,
                "exported": [{
                    "vendor_key": "preview",
                    "source_path": str(RESMAN_TEMPLATE),
                    "export_path": str(dest),
                    "filename": dest.name,
                }],
                "export_used_edited_rows": True,
                "export_used_preview_cache": True,
                "edited_rows_count": len(preview_rows),
                "rows_written": rows_written,
                "document_url_updates": document_url_updates,
                "document_url_warnings": document_url_warnings,
                "accounting_readiness": readiness,
            }
        return {"batch_id": batch_id, "exported": [], "reason": "no_processed_output_yet",
                "export_used_edited_rows": False}

    # Legacy workbooks may predate the row contract and cannot prove invoice
    # reconciliation. Rebuild from cached rows through the same gate instead
    # of copying an opaque workbook.
    exported: list[dict] = []
    preview_rows = _cached_preview_rows_for_export(batch_id)
    if preview_rows:
            _normalize_edited_rows_for_export(batch_id, preview_rows)
            document_url_warnings = _ensure_dropbox_document_urls_for_export(batch_id, preview_rows)
            _persist_export_document_urls_to_cache(batch_id, preview_rows)
            document_url_updates = _document_url_updates_for_rows(batch_id, preview_rows)
            from .validated_export_bridge import ExportAuthorizationError, ReadinessValidatedExporter
            try:
                readiness = ReadinessValidatedExporter().authorize(batch_id, preview_rows)
            except ExportAuthorizationError as exc:
                return {"batch_id": batch_id, "exported": [], "reason": exc.code, "blockers": exc.blockers}
            _validate_resman_export_required_rows(preview_rows)
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            dest = export_dir / f"resman_import_preview_{ts}.xlsx"
            rows_written = _write_edited_rows_to_template(RESMAN_TEMPLATE, dest, preview_rows)
            return {
                "batch_id": batch_id,
                "exported": [{
                    "vendor_key": "preview",
                    "source_path": str(RESMAN_TEMPLATE),
                    "export_path": str(dest),
                    "filename": dest.name,
                }],
                "export_used_edited_rows": True,
                "export_used_preview_cache": True,
                "edited_rows_count": len(preview_rows),
                "rows_written": rows_written,
                "document_url_updates": document_url_updates,
                "document_url_warnings": document_url_warnings,
                "accounting_readiness": readiness,
            }

    if any(processed_dir.glob("*/**/*_resman_import_*.xlsx")):
        return {"batch_id": batch_id, "exported": [], "reason": "legacy_export_disabled",
                "export_used_edited_rows": False}

    return {
        "batch_id": batch_id,
        "exported": exported,
        "export_used_edited_rows": False,
    }
